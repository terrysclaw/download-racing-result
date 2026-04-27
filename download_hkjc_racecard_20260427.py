import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import date
import numpy as np
import logging
from typing import Dict, List, Optional, Tuple
from pathlib import Path
import matplotlib.pyplot as plt
from matplotlib import rcParams
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class HKJCRacecardDownloader:
    def __init__(self, race_date: date, race_course: str):
        self.race_date = race_date
        self.race_course = race_course
        self.base_url = "https://racing.hkjc.com/racing/information/Chinese/Racing"
        self.df_results = self._load_historical_data()
        
    def _load_historical_data(self) -> pd.DataFrame:
        """Load historical race results."""
        try:
            df_2024 = pd.read_excel('raceresult_2024.xlsx')
            df_2025 = pd.read_excel('raceresult_2025.xlsx')
            return pd.concat([df_2024, df_2025], ignore_index=True)
        except FileNotFoundError as e:
            logger.error(f"Historical data file not found: {e}")
            return pd.DataFrame()
    
    def _initialize_data_structure(self) -> Dict[str, List]:
        """Initialize the data structure for racecard."""
        return {
            '日期': [], '馬場': [], '場次': [], '泥草': [], '賽道': [], '班次': [], '路程': [],
            '馬匹編號': [], '6次近績': [], '馬名': [], '網址': [], '烙號': [], '負磅': [],
            '騎師': [], '可能超磅': [], '檔位': [], '練馬師': [], '國際評分': [], '評分': [],
            '評分+/-': [], '排位體重': [], '排位體重+/-': [], '最佳時間': [], '馬齡': [],
            '分齡讓磅': [], '性別': [], '今季獎金': [], '優先參賽次序': [], '上賽距今日數': [],
            '配備': [], '馬主': [], '父系': [], '母系': [], '進口類別': [],
        }
    
    def _safe_int_convert(self, value: str) -> Optional[int]:
        """Safely convert string to integer."""
        try:
            return int(value.strip())
        except (ValueError, AttributeError):
            return None
    
    def _clean_text(self, text: str) -> str:
        """Clean text by removing unwanted characters."""
        return text.replace('\r', '').replace('\n', '').replace('\xa0', '').strip()
    
    def _extract_race_info(self, soup: BeautifulSoup) -> Tuple[Optional[str], Optional[str], Optional[int]]:
        """Extract race information from soup."""
        race_div = soup.find('div', class_='f_fs13')
        
        if not race_div:
            return None, None, None
            
        band = None
        race_class = None
        distance = None
        
        for col in race_div.text.split(','):
            if '賽道' in col:
                band = col.strip()
            if '班' in col:
                race_class = col.strip()
            if '米' in col:
                try:
                    distance = self._safe_int_convert(col.split('米')[0])
                except:
                    distance = None
                    
        return band, race_class, distance
    
    def _extract_horse_data(self, cells: List, race_info: Dict, data: Dict[str, List]) -> None:
        """Extract individual horse data from table cells."""
        # Add race information
        for key, value in race_info.items():
            data[key].append(value)
        
        # Extract horse-specific data with safe conversion
        data['馬匹編號'].append(self._safe_int_convert(cells[0].text))
        data['6次近績'].append(self._clean_text(cells[1].text) if len(cells) > 1 else None)
        
        # Horse name and URL
        if len(cells) > 3:
            data['馬名'].append(self._clean_text(cells[3].text))
            try:
                data['網址'].append("https://racing.hkjc.com" + cells[3].find('a')['href'])
            except:
                data['網址'].append(None)
        else:
            data['馬名'].append(None)
            data['網址'].append(None)
        
        # Continue with other fields
        fields_mapping = [
            ('烙號', 4), ('負磅', 5), ('騎師', 6), ('可能超磅', 7), ('檔位', 8),
            ('練馬師', 9), ('國際評分', 10), ('評分', 11), ('評分+/-', 12),
            ('排位體重', 13), ('排位體重+/-', 14), ('最佳時間', 15), ('馬齡', 16),
            ('分齡讓磅', 17), ('性別', 18), ('今季獎金', 19), ('優先參賽次序', 20),
            ('上賽距今日數', 21), ('配備', 22), ('馬主', 23), ('父系', 24),
            ('母系', 25), ('進口類別', 26)
        ]
        
        for field_name, index in fields_mapping:
            if index < len(cells):
                if field_name in ['負磅', '檔位', '國際評分', '評分', '排位體重', '馬齡', '分齡讓磅']:
                    data[field_name].append(self._safe_int_convert(cells[index].text))
                else:
                    data[field_name].append(self._clean_text(cells[index].text))
            else:
                data[field_name].append(None)
    
    def _get_time_adjustment_factors(self, venue: str, surface: str, distance: int) -> Dict[str, float]:
        """Get time adjustment factors based on venue, surface, and distance."""
        factors = {'weight_early': 0, 'weight_late': 0, 'gate': 0}
        
        if venue == 'HV':
            distance_factors = {
                1000: {'weight_early': 0.015, 'weight_late': 0.056, 'gate': 0.02},
                1200: {'weight_early': 0.017, 'weight_late': 0.056, 'gate': 0.033},
                1650: {'weight_early': 0.019, 'weight_late': 0.056, 'gate': 0.043},
                1800: {'weight_early': 0.021, 'weight_late': 0.056, 'gate': 0.048},
                2200: {'weight_early': 0.023, 'weight_late': 0.056, 'gate': 0.053}
            }
        else:  # ST
            if surface == '草地':
                distance_factors = {
                    1000: {'weight_early': 0.015, 'weight_late': 0.056, 'gate': -0.019},
                    1200: {'weight_early': 0.017, 'weight_late': 0.056, 'gate': 0.028},
                    1400: {'weight_early': 0.018, 'weight_late': 0.056, 'gate': 0.033},
                    1600: {'weight_early': 0.019, 'weight_late': 0.056, 'gate': 0.038},
                    1800: {'weight_early': 0.021, 'weight_late': 0.056, 'gate': 0.043},
                    2000: {'weight_early': 0.022, 'weight_late': 0.056, 'gate': 0.048},
                    2400: {'weight_early': 0.024, 'weight_late': 0.056, 'gate': 0.053}
                }
            else:  # 泥地
                distance_factors = {
                    1200: {'weight_early': 0.017, 'weight_late': 0.056, 'gate': 0.028},
                    1650: {'weight_early': 0.019, 'weight_late': 0.056, 'gate': 0.043},
                    1800: {'weight_early': 0.021, 'weight_late': 0.056, 'gate': 0.043}
                }
        
        return distance_factors.get(distance, factors)
    
    def _to_float(self, value) -> Optional[float]:
        """Safely convert mixed value to float."""
        if value is None or pd.isna(value) or value == '' or value == '---':
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _get_section_count_by_distance(self, distance_value) -> int:
        """Map race distance to practical HK sectional count."""
        distance = self._to_float(distance_value)
        if distance is None:
            return 6
        if distance <= 1200:
            return 3
        if distance <= 1650:
            return 4
        if distance <= 2000:
            return 5
        return 6

    def _get_recent_three_runs(self, row: pd.Series) -> pd.DataFrame:
        """Get recent 3 runs for sectional forecasting (same venue required)."""
        if self.df_results.empty:
            return pd.DataFrame()

        same_venue_filter = (
            (self.df_results['布號'] == row['烙號']) &
            (self.df_results['馬場'] == row['馬場']) &
            (self.df_results['泥草'] == row['泥草']) &
            (self.df_results['路程'] == row['路程']) &
            (self.df_results['獨贏賠率'] > 0)
        )
        return self.df_results[same_venue_filter].tail(3)

    def _calculate_predicted_sections(
        self,
        row: pd.Series,
        avg_sections: Dict[int, Optional[float]],
        recent_runs: pd.DataFrame,
        target_section_count: int,
        jt_rate: Optional[float] = None
    ) -> Dict[int, Optional[float]]:
        """Estimate today's sectional times using weight, draw, and jockey-trainer adjustments."""
        predicted = {i: (avg_sections.get(i) if i <= target_section_count else None) for i in range(1, 7)}

        valid_section_count = len([avg_sections.get(i) for i in range(1, target_section_count + 1) if avg_sections.get(i) is not None])
        if valid_section_count == 0:
            return predicted

        distance = self._to_float(row.get('路程'))
        if not distance or distance <= 0:
            return predicted

        current_weight = self._to_float(row.get('負磅'))
        current_gate = self._to_float(row.get('檔位'))
        hist_weight_mean = pd.to_numeric(recent_runs.get('實際負磅'), errors='coerce').mean()
        hist_gate_mean = pd.to_numeric(recent_runs.get('檔位'), errors='coerce').mean()

        weight_diff = 0.0
        gate_diff = 0.0
        if current_weight is not None and pd.notna(hist_weight_mean):
            weight_diff = current_weight - float(hist_weight_mean)
        if current_gate is not None and pd.notna(hist_gate_mean):
            gate_diff = current_gate - float(hist_gate_mean)

        factors = self._get_time_adjustment_factors(
            row['馬場'], row['泥草'], int(distance)
        )

        section_distance = distance / valid_section_count
        early_adj_per_section = weight_diff * factors['weight_early'] / distance * section_distance
        late_adj_per_section = weight_diff * factors['weight_late'] / distance * section_distance

        for i in range(1, target_section_count + 1):
            base_time = avg_sections.get(i)
            if base_time is None:
                continue

            # Gate impact is concentrated in the first 2 sections.
            if i == 1:
                gate_adj = gate_diff * factors['gate'] * 0.7
            elif i == 2:
                gate_adj = gate_diff * factors['gate'] * 0.3
            else:
                gate_adj = 0.0

            half_point = max(1, valid_section_count // 2)
            weight_adj = early_adj_per_section if i <= half_point else late_adj_per_section
            predicted[i] = round(base_time + weight_adj + gate_adj, 3)

        # Jockey-trainer adjustment: pairs above/below 33% baseline run proportionally faster/slower.
        # Coefficient 0.02 means a 50% pair (~+17% above baseline) adjusts each section by ~-0.34%.
        if jt_rate is not None:
            jt_factor = 1.0 - (jt_rate - 0.33) * 0.02
            for i in range(1, target_section_count + 1):
                if predicted[i] is not None:
                    predicted[i] = round(predicted[i] * jt_factor, 3)

        return predicted

    def _simulate_race_shape(self, df: pd.DataFrame) -> pd.DataFrame:
        """Simulate likely race positioning from forecast early pace and draw."""
        if df.empty or '預估頭段(第1+2段)' not in df.columns:
            return df

        df['預計各分段排名'] = None
        df['模擬走勢'] = None
        df['模擬走位'] = None

        for race_no, race_df in df.groupby('場次'):
            valid_index = race_df[race_df['預估頭段(第1+2段)'].notna()].index
            if len(valid_index) == 0:
                continue

            # Build per-section cumulative ranking sequence, e.g. 2-2-1-1.
            rank_sequences = {idx: [] for idx in valid_index}
            cumulative_times = pd.Series(0.0, index=valid_index, dtype='float64')

            for section_no in range(1, 7):
                section_col = f'預估今場第 {section_no} 段'
                section_values = pd.to_numeric(df.loc[valid_index, section_col], errors='coerce')

                valid_mask = section_values.notna() & cumulative_times.notna()
                if not valid_mask.any():
                    continue

                cumulative_times.loc[valid_mask] = cumulative_times.loc[valid_mask] + section_values.loc[valid_mask]
                cumulative_times.loc[~valid_mask] = np.nan

                section_rank = cumulative_times.rank(method='min', ascending=True)
                for idx in valid_index:
                    rank_value = section_rank.loc[idx]
                    if pd.notna(rank_value):
                        rank_sequences[idx].append(str(int(rank_value)))

            for idx in valid_index:
                if rank_sequences[idx]:
                    df.at[idx, '預計各分段排名'] = '-'.join(rank_sequences[idx])

            early_rank = df.loc[valid_index, '預估頭段(第1+2段)'].rank(method='min', ascending=True)
            field_size = len(valid_index)

            for idx in valid_index:
                rank_value = int(early_rank.loc[idx])
                gate = self._to_float(df.at[idx, '檔位'])

                if rank_value <= max(1, int(np.ceil(field_size * 0.3))):
                    pace_bucket = '前置'
                elif rank_value <= max(2, int(np.ceil(field_size * 0.7))):
                    pace_bucket = '中置'
                else:
                    pace_bucket = '後上'

                if gate is None:
                    gate_bucket = '不明檔位'
                elif gate <= 4:
                    gate_bucket = '內檔'
                elif gate <= 8:
                    gate_bucket = '中檔'
                else:
                    gate_bucket = '外檔'

                if pace_bucket == '前置' and gate_bucket == '內檔':
                    style = '前置貼欄'
                elif pace_bucket == '前置' and gate_bucket == '外檔':
                    style = '外疊搶放'
                elif pace_bucket == '中置' and gate_bucket == '內檔':
                    style = '包廂跟前'
                elif pace_bucket == '後上' and gate_bucket == '外檔':
                    style = '留後外疊'
                elif pace_bucket == '後上':
                    style = '留後追趕'
                else:
                    style = f'{pace_bucket}守位'

                df.at[idx, '模擬走勢'] = style
                df.at[idx, '模擬走位'] = f"{style} ({gate_bucket}, 頭段排名{rank_value})"

        return df

    def _get_jockey_trainer_pair(self, jockey: str, trainer: str) -> Optional[pd.DataFrame]:
        """Return historical rows for a jockey-trainer pair, stripping overweight notation."""
        import re
        jockey = re.sub(r'\s*\([^)]*\)', '', str(jockey or '')).strip()
        trainer = re.sub(r'\s*\([^)]*\)', '', str(trainer or '')).strip()
        if self.df_results.empty or not jockey or not trainer:
            return None
        pair = self.df_results[
            (self.df_results['騎師'] == jockey) &
            (self.df_results['練馬師'] == trainer)
        ]
        return pair if not pair.empty else None

    def _calculate_jockey_trainer_rate(self, jockey: str, trainer: str) -> Optional[str]:
        """Return jockey-trainer place rate (top 3) as formatted string."""
        pair = self._get_jockey_trainer_pair(jockey, trainer)
        if pair is None:
            return None
        total = len(pair)
        placed = pair['名次'].apply(lambda x: pd.to_numeric(x, errors='coerce')).le(3).sum()
        rate = placed / total
        return f"{rate:.0%}({int(placed)}/{total})"

    def _calculate_jockey_trainer_rate_numeric(self, jockey: str, trainer: str) -> Optional[float]:
        """Return jockey-trainer place rate (top 3) as float (0.0–1.0), or None."""
        pair = self._get_jockey_trainer_pair(jockey, trainer)
        if pair is None:
            return None
        total = len(pair)
        placed = pair['名次'].apply(lambda x: pd.to_numeric(x, errors='coerce')).le(3).sum()
        return placed / total

    def _add_sectional_forecast(self, df: pd.DataFrame) -> pd.DataFrame:
        """Add 3-run sectional averages and forecast sectionals for today's race."""
        avg_cols = [f'近3仗第 {i} 段平均' for i in range(1, 7)]
        pred_cols = [f'預估今場第 {i} 段' for i in range(1, 7)]
        extra_cols = ['近3仗樣本數', '預估頭段(第1+2段)', '預估末段(最後2段)', '預估末段(第5+6段)', '預估總段速', '預計各分段排名', '模擬走勢', '模擬走位', '騎練合作上名率']

        for col in avg_cols + pred_cols + extra_cols:
            df[col] = None

        for index, row in df.iterrows():
            recent_runs = self._get_recent_three_runs(row)
            if recent_runs.empty:
                continue

            target_section_count = self._get_section_count_by_distance(row.get('路程'))
            sample_count = len(recent_runs)
            avg_sections = {}
            for i in range(1, 7):
                if i <= target_section_count:
                    section_series = pd.to_numeric(recent_runs.get(f'第 {i} 段'), errors='coerce')
                    avg_value = section_series.mean()
                    avg_sections[i] = round(float(avg_value), 3) if pd.notna(avg_value) else None
                else:
                    avg_sections[i] = None
                df.at[index, f'近3仗第 {i} 段平均'] = avg_sections[i]

            jt_rate = self._calculate_jockey_trainer_rate_numeric(
                str(row.get('騎師', '') or '').strip(),
                str(row.get('練馬師', '') or '').strip()
            )
            predicted_sections = self._calculate_predicted_sections(
                row,
                avg_sections,
                recent_runs,
                target_section_count,
                jt_rate=jt_rate
            )
            for i in range(1, 7):
                df.at[index, f'預估今場第 {i} 段'] = predicted_sections.get(i)

            head = [predicted_sections.get(1), predicted_sections.get(2)]
            tail_start = max(1, target_section_count - 1)
            tail = [predicted_sections.get(i) for i in range(tail_start, target_section_count + 1)]
            valid_head = [v for v in head if v is not None]
            valid_tail = [v for v in tail if v is not None]
            valid_all = [predicted_sections.get(i) for i in range(1, target_section_count + 1) if predicted_sections.get(i) is not None]

            df.at[index, '近3仗樣本數'] = sample_count
            df.at[index, '預估頭段(第1+2段)'] = round(sum(valid_head), 3) if valid_head else None
            df.at[index, '預估末段(最後2段)'] = round(sum(valid_tail), 3) if valid_tail else None
            df.at[index, '預估末段(第5+6段)'] = df.at[index, '預估末段(最後2段)']
            df.at[index, '預估總段速'] = round(sum(valid_all), 3) if valid_all else None

        df['騎練合作上名率'] = df.apply(
            lambda row: self._calculate_jockey_trainer_rate(
                str(row.get('騎師', '') or '').strip(),
                str(row.get('練馬師', '') or '').strip()
            ), axis=1
        )

        return self._simulate_race_shape(df)
    
    def _scrape_race_data(self, race_no: int) -> Optional[pd.DataFrame]:
        """Scrape race data for a specific race."""
        url = f"{self.base_url}/RaceCard.aspx?RaceDate={self.race_date.strftime('%Y/%m/%d')}&Racecourse={self.race_course}&RaceNo={race_no}"
        
        try:
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Extract race information
            band, race_class, distance = self._extract_race_info(soup)
            
            race_info = {
                '日期': self.race_date,
                '馬場': self.race_course,
                '場次': race_no,
                '泥草': '草地' if band else '泥地',
                '賽道': band,
                '班次': race_class,
                '路程': distance
            }
            
            # Find racecard table
            racecard_table = soup.find('table', id='racecardlist')
            if not racecard_table:
                logger.warning(f"No racecard table found for Race {race_no}")
                return None
            
            # Extract horse data
            data = self._initialize_data_structure()
            tbody = racecard_table.find('tbody')
            
            if tbody:
                for tr in tbody.find_all('tr')[2:]:  # Skip header rows
                    cells = tr.find_all('td')
                    if len(cells) >= 4:  # Ensure minimum required cells
                        self._extract_horse_data(cells, race_info, data)
            
            df = pd.DataFrame(data)
            
            if not df.empty:
                df = self._add_sectional_forecast(df)
            
            return df
            
        except requests.RequestException as e:
            logger.error(f"Error fetching race {race_no}: {e}")
            return None
        except Exception as e:
            logger.error(f"Unexpected error processing race {race_no}: {e}")
            return None
    
    def _create_terry_sheet(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create Terry AI formatted sheet."""
        required_columns = [
            '場次', '班次', '路程', '馬匹編號', '馬名', '騎師', '練馬師', '騎練合作上名率', '評分',
            '負磅', '檔位', '最佳時間',
            '近3仗樣本數',
            '近3仗第 1 段平均', '近3仗第 2 段平均', '近3仗第 3 段平均',
            '近3仗第 4 段平均', '近3仗第 5 段平均', '近3仗第 6 段平均',
            '預估今場第 1 段', '預估今場第 2 段', '預估今場第 3 段',
            '預估今場第 4 段', '預估今場第 5 段', '預估今場第 6 段',
            '預估頭段(第1+2段)', '預估末段(最後2段)', '預估末段(第5+6段)', '預估總段速', '預計各分段排名', '模擬走勢', '模擬走位',
            '網址'
        ]

        existing_columns = [col for col in required_columns if col in df.columns]
        df_terry = df[existing_columns].copy()
        return df_terry
    
    def _create_combined_files(self, date_str: str) -> None:
        """Create combined Excel files for different formats."""
        def sort_by_predicted_total(df: pd.DataFrame) -> pd.DataFrame:
            """Sort by 預估總段速 ascending (faster first), keeping missing values at the end."""
            if df.empty or '預估總段速' not in df.columns:
                return df

            sorted_df = df.copy()
            sorted_df['預估總段速數值'] = pd.to_numeric(sorted_df['預估總段速'], errors='coerce')

            sort_cols = ['預估總段速數值']
            ascending = [True]

            # Stable tie-breakers for readability.
            if '檔位' in sorted_df.columns:
                sorted_df['檔位數值'] = pd.to_numeric(sorted_df['檔位'], errors='coerce')
                sort_cols.append('檔位數值')
                ascending.append(True)
            elif '馬匹編號' in sorted_df.columns:
                sorted_df['馬匹編號數值'] = pd.to_numeric(sorted_df['馬匹編號'], errors='coerce')
                sort_cols.append('馬匹編號數值')
                ascending.append(True)

            sorted_df = sorted_df.sort_values(by=sort_cols, ascending=ascending, na_position='last')

            drop_cols = ['預估總段速數值', '檔位數值', '馬匹編號數值']
            existing_drop_cols = [col for col in drop_cols if col in sorted_df.columns]
            return sorted_df.drop(columns=existing_drop_cols)

        # Standard combined file
        standard_sheets_created = 0
        with pd.ExcelWriter(f"Racecard_{date_str}.xlsx", engine='openpyxl') as writer:
            for race_no in range(1, 12):
                try:
                    df = pd.read_excel(f"Racecard_{date_str}_{race_no}.xlsx")
                    if not df.empty:
                        df = sort_by_predicted_total(df)
                        df.to_excel(writer, sheet_name=f"Race_{race_no}", index=False)
                        standard_sheets_created += 1
                except FileNotFoundError:
                    logger.warning(f"File Racecard_{date_str}_{race_no}.xlsx not found")
                except Exception as e:
                    logger.error(f"Error adding Race {race_no} to standard file: {e}")
        
        if standard_sheets_created == 0:
            logger.warning(f"No sheets created for standard file Racecard_{date_str}.xlsx")
        
        # Terry format
        terry_sheets_created = 0
        with pd.ExcelWriter(f"Racecard_{date_str}_Terry.xlsx", engine='openpyxl') as writer:
            for race_no in range(1, 12):
                try:
                    df = pd.read_excel(f"Racecard_{date_str}_{race_no}.xlsx")
                    if not df.empty:
                        df = sort_by_predicted_total(df)
                        df_terry = self._create_terry_sheet(df)
                        if not df_terry.empty:
                            df_terry.to_excel(writer, sheet_name=f"Race_{race_no}", index=False)
                            terry_sheets_created += 1
                except FileNotFoundError:
                    logger.warning(f"File Racecard_{date_str}_{race_no}.xlsx not found for Terry format")
                except Exception as e:
                    logger.error(f"Error creating Terry sheet for Race {race_no}: {e}")
            
            # Create a dummy sheet if no sheets were created
            if terry_sheets_created == 0:
                dummy_df = pd.DataFrame({'Message': ['No race data available']})
                dummy_df.to_excel(writer, sheet_name="No_Data", index=False)
                logger.warning("Created dummy sheet for Terry format due to no valid data")
        
        logger.info(f"Created combined files: Standard({standard_sheets_created} sheets), Terry({terry_sheets_created} sheets)")

    def _generate_positioning_plot(self, df: pd.DataFrame, race_no: int, date_str: str) -> None:
        """Generate a visualization plot for race positioning."""
        try:
            rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans']
            rcParams['axes.unicode_minus'] = False
            
            valid_data = df[
                (df['檔位'].notna())
            ].copy()
            
            if valid_data.empty:
                return

            valid_data['檔位數值'] = pd.to_numeric(valid_data['檔位'], errors='coerce')
            valid_data = valid_data.dropna(subset=['檔位數值']).copy()
            if valid_data.empty:
                return

            # 頭400: 第1+2段合計 (early pace)
            valid_data['第1段預估'] = pd.to_numeric(valid_data.get('預估頭段(第1+2段)'), errors='coerce')

            # 終點: 預估總段速 (overall total)
            valid_data['最終預計'] = pd.to_numeric(valid_data.get('預估總段速'), errors='coerce')

            # Keep horses with missing metrics by pushing them to the back of the field.
            def rank_with_missing_last(metric_col: str) -> pd.Series:
                metric = valid_data[metric_col]
                rank = metric.rank(method='min', ascending=True)
                missing_mask = metric.isna()
                if missing_mask.any():
                    start_rank = int(rank.max()) + 1 if rank.notna().any() else 1
                    fallback_order = valid_data.loc[missing_mask, '檔位數值'].rank(method='first', ascending=True).astype(int)
                    rank.loc[missing_mask] = start_rank + fallback_order - 1
                return rank

            valid_data['第1段排名'] = rank_with_missing_last('第1段預估')
            valid_data['最終排名'] = rank_with_missing_last('最終預計')

            fig, axes = plt.subplots(2, 1, figsize=(13, 7.5))
            fig.patch.set_facecolor('#D9D9D9')

            def draw_stage(stage_ax, rank_col: str, stage_title: str) -> None:
                stage_ax.set_facecolor('#D9D9D9')
                stage_df = valid_data.dropna(subset=[rank_col]).sort_values(by=[rank_col, '檔位數值']).copy()
                if stage_df.empty:
                    stage_ax.text(0.5, 0.5, f'{stage_title}: 無數據', ha='center', va='center', fontsize=13, color='#666666')
                    stage_ax.set_xticks([])
                    stage_ax.set_yticks([])
                    for spine in stage_ax.spines.values():
                        spine.set_visible(False)
                    return

                horse_count = len(stage_df)
                front_cut = max(1, int(np.ceil(horse_count * 0.2)))
                mid_cut = max(front_cut + 1, int(np.ceil(horse_count * 0.55)))
                row_levels = []
                for seq_no in range(1, horse_count + 1):
                    if seq_no <= front_cut:
                        row_levels.append(3.0)
                    elif seq_no <= mid_cut:
                        row_levels.append(2.0)
                    else:
                        row_levels.append(1.0)
                stage_df['row_level'] = row_levels

                # Re-layout by stage ranking so three panels show clear differences.
                stage_df['display_x'] = np.nan
                row_spans = {
                    3.0: (4.5, 7.5),
                    2.0: (2.5, 9.5),
                    1.0: (1.0, 11.0)
                }
                for level, (x_min, x_max) in row_spans.items():
                    row_mask = stage_df['row_level'] == level
                    row_count = int(row_mask.sum())
                    if row_count == 0:
                        continue
                    if row_count == 1:
                        x_positions = np.array([(x_min + x_max) / 2])
                    else:
                        x_positions = np.linspace(x_min, x_max, row_count)
                    stage_df.loc[row_mask, 'display_x'] = x_positions

                for _, stage_row in stage_df.iterrows():
                    x = float(stage_row['display_x'])
                    y = float(stage_row['row_level'])
                    horse_no = str(stage_row.get('馬匹編號', '')).strip()
                    gate_no = str(stage_row.get('檔位', '')).strip()

                    stage_ax.scatter(
                        x,
                        y,
                        s=520,
                        marker='v',
                        color='#C2410C',
                        edgecolors='#9A3412',
                        linewidth=1.2,
                        zorder=2
                    )
                    stage_ax.text(
                        x,
                        y + 0.03,
                        horse_no,
                        ha='center',
                        va='center',
                        fontsize=11,
                        color='white',
                        fontweight='bold',
                        zorder=3
                    )
                    if gate_no:
                        stage_ax.text(
                            x,
                            y - 0.28,
                            f'檔{gate_no}',
                            ha='center',
                            va='center',
                            fontsize=8,
                            color='#6B7280',
                            zorder=3
                        )

                stage_ax.set_xlim(0.3, 11.7)
                stage_ax.set_ylim(0.25, 3.55)
                stage_ax.set_xticks([])
                stage_ax.set_yticks([])
                stage_ax.set_title(stage_title, loc='left', fontsize=12, color='#333333', pad=6)
                for spine in stage_ax.spines.values():
                    spine.set_visible(False)

            draw_stage(axes[0], '最終排名', '終點')
            draw_stage(axes[1], '第1段排名', '頭800')

            fig.suptitle(f'{date_str} Race {race_no} 走位模擬', fontsize=14, color='#222222', y=0.99)
            plt.tight_layout(rect=[0, 0, 1, 0.97])
            
            # Save figure
            output_path = f"Racecard_{date_str}_走位模擬_Race{race_no:02d}.png"
            plt.savefig(output_path, dpi=100, bbox_inches='tight')
            plt.close()
            
            logger.info(f"Generated positioning plot: {output_path}")
        except Exception as e:
            logger.error(f"Error generating positioning plot for Race {race_no}: {e}")

    def _create_positioning_file(self, date_str: str) -> None:
        """Create a dedicated Excel file for simulated running positions and generate visualization plots."""
        positioning_columns = [
            '場次', '馬匹編號', '馬名', '評分', '負磅', '檔位', '騎師', '練馬師', '騎練合作上名率',
            '近3仗樣本數',
            '預估今場第 1 段', '預估今場第 2 段', '預估今場第 3 段',
            '預估今場第 4 段', '預估今場第 5 段', '預估今場第 6 段',
            '預估頭段(第1+2段)', '預估末段(最後2段)', '預估總段速',
            '預計各分段排名', '模擬走勢', '模擬走位'
        ]

        def build_prediction_dataframe(prediction_map: Dict[int, pd.DataFrame]) -> pd.DataFrame:
            """Build prediction table with one row per race and only top 4 by 預估總段速."""
            if not prediction_map:
                return pd.DataFrame()

            race_nos = sorted(prediction_map.keys())
            rows: List[Dict[str, str]] = []

            for race_no in race_nos:
                race_df = prediction_map[race_no].head(4).copy()
                horse_nos = race_df['馬匹編號'].astype(str).tolist()
                horse_names = race_df['馬名'].fillna('').astype(str).tolist()
                if '預估總段速數值' in race_df.columns:
                    total_times = pd.to_numeric(race_df['預估總段速數值'], errors='coerce').tolist()
                elif '預估總段速' in race_df.columns:
                    total_times = pd.to_numeric(race_df['預估總段速'], errors='coerce').tolist()
                else:
                    total_times = []

                while len(horse_nos) < 4:
                    horse_nos.append('')
                    horse_names.append('')
                while len(total_times) < 2:
                    total_times.append(np.nan)

                if pd.notna(total_times[0]) and pd.notna(total_times[1]):
                    top2_gap = round(float(total_times[1]) - float(total_times[0]), 3)
                else:
                    top2_gap = ''

                row: Dict[str, str] = {'場次': f'第{race_no}場'}
                for rank_idx in range(4):
                    row[f'第{rank_idx + 1}名馬號'] = horse_nos[rank_idx]
                    row[f'第{rank_idx + 1}名馬名'] = horse_names[rank_idx]
                row['第1-2名預估總段速差距'] = top2_gap
                rows.append(row)

            return pd.DataFrame(rows)

        def style_prediction_sheet(writer: pd.ExcelWriter, max_row: int, max_col: int) -> None:
            """Apply reference-like formatting (yellow/gray blocks) for prediction sheet."""
            ws = writer.book['prediction']

            yellow_fill = PatternFill(fill_type='solid', fgColor='FFC000')
            gray_fill = PatternFill(fill_type='solid', fgColor='E6E6E6')
            header_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
            thin_side = Side(style='thin', color='BFBFBF')
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

            # Header style
            for col in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # Body style: race gray; horse no yellow; horse name gray.
            for row in range(2, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if col == 1:
                        cell.fill = gray_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col == max_col:
                        cell.fill = gray_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col % 2 == 0:
                        cell.fill = yellow_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.fill = gray_fill
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = thin_border

            ws.freeze_panes = 'A2'
            ws.column_dimensions['A'].width = 6
            for col in range(2, max_col + 1):
                col_letter = ws.cell(row=1, column=col).column_letter
                if col % 2 == 0:
                    ws.column_dimensions[col_letter].width = 10
                else:
                    ws.column_dimensions[col_letter].width = 14

        sheets_created = 0
        plots_created = 0
        prediction_data: Dict[int, pd.DataFrame] = {}
        with pd.ExcelWriter(f"Racecard_{date_str}_走位模擬.xlsx", engine='openpyxl') as writer:
            for race_no in range(1, 12):
                try:
                    df = pd.read_excel(f"Racecard_{date_str}_{race_no}.xlsx")
                    if df.empty:
                        continue

                    existing_columns = [col for col in positioning_columns if col in df.columns]
                    if not existing_columns:
                        continue

                    df_positioning = df[existing_columns].copy()
                    if '預估總段速' in df_positioning.columns:
                        df_positioning['預估總段速數值'] = pd.to_numeric(df_positioning['預估總段速'], errors='coerce')
                        if '檔位' in df_positioning.columns:
                            df_positioning['檔位數值'] = pd.to_numeric(df_positioning['檔位'], errors='coerce')
                            df_positioning = df_positioning.sort_values(
                                by=['預估總段速數值', '檔位數值'],
                                ascending=[True, True],
                                na_position='last'
                            )
                            df_positioning = df_positioning.drop(columns=['檔位數值'])
                        else:
                            df_positioning = df_positioning.sort_values(
                                by=['預估總段速數值'],
                                ascending=[True],
                                na_position='last'
                            )
                        df_positioning = df_positioning.drop(columns=['預估總段速數值'])

                    # Build prediction source sorted by 預估總段速 (smaller = faster).
                    if {'馬匹編號', '馬名', '預估總段速'}.issubset(df.columns):
                        pred_df = df[['馬匹編號', '馬名', '檔位', '預估總段速']].copy()
                        pred_df['檔位數值'] = pd.to_numeric(pred_df.get('檔位'), errors='coerce')
                        pred_df['預估總段速數值'] = pd.to_numeric(pred_df['預估總段速'], errors='coerce')
                        pred_df['缺值'] = pred_df['預估總段速數值'].isna()
                        pred_df = pred_df.sort_values(
                            by=['缺值', '預估總段速數值', '檔位數值'],
                            ascending=[True, True, True],
                            na_position='last'
                        )
                        prediction_data[race_no] = pred_df[['馬匹編號', '馬名', '預估總段速數值']].reset_index(drop=True)

                    df_positioning.to_excel(writer, sheet_name=f"Race_{race_no}", index=False)
                    sheets_created += 1
                    
                    # Generate positioning visualization plot
                    self._generate_positioning_plot(df, race_no, date_str)
                    plots_created += 1
                except FileNotFoundError:
                    logger.warning(f"File Racecard_{date_str}_{race_no}.xlsx not found for positioning format")
                except Exception as e:
                    logger.error(f"Error creating positioning sheet for Race {race_no}: {e}")

            if sheets_created == 0:
                dummy_df = pd.DataFrame({'Message': ['No race positioning data available']})
                dummy_df.to_excel(writer, sheet_name="No_Data", index=False)
                logger.warning("Created dummy sheet for positioning format due to no valid data")
            elif prediction_data:
                prediction_sheet_df = build_prediction_dataframe(prediction_data)
                prediction_sheet_df.to_excel(writer, sheet_name='prediction', index=False)
                style_prediction_sheet(
                    writer,
                    max_row=prediction_sheet_df.shape[0] + 1,
                    max_col=prediction_sheet_df.shape[1]
                )

        logger.info(f"Created positioning file: Racecard_{date_str}_走位模擬.xlsx ({sheets_created} sheets) and {plots_created} visualization plots")
    
    def download_all_racecards(self) -> None:
        """Download all racecards for the configured race date and course."""
        date_str = self.race_date.strftime('%Y%m%d')
        logger.info(f"Starting racecard download for {self.race_date} at {self.race_course}")
        
        successful_downloads = 0
        individual_files = []  # Track individual files created
        
        # Download individual race cards (typically 1-11 races)
        for race_no in range(1, 12):
            logger.info(f"Processing {self.race_date} {self.race_course} Race {race_no}")
            
            df = self._scrape_race_data(race_no)
            
            if df is not None and not df.empty:
                # Save individual race file
                filename = f"Racecard_{date_str}_{race_no}.xlsx"
                try:
                    df.to_excel(filename, index=False)
                    logger.info(f"Saved {filename}")
                    successful_downloads += 1
                    individual_files.append(filename)  # Track the file
                except Exception as e:
                    logger.error(f"Error saving {filename}: {e}")
            else:
                logger.warning(f"No data found for Race {race_no}")
        
        # Create combined files in different formats
        if successful_downloads > 0:
            try:
                self._create_combined_files(date_str)
                self._create_positioning_file(date_str)
                
                # Remove individual race files after successful combined file creation
                self._cleanup_individual_files(individual_files)
                
            except Exception as e:
                logger.error(f"Error creating combined files: {e}")
        else:
            logger.warning("No successful downloads, skipping combined file creation")
        
        logger.info(f"Racecard download completed. Successfully processed {successful_downloads} races.")
    
    def _cleanup_individual_files(self, individual_files: List[str]) -> None:
        """Remove individual race files after combined files are created."""
        removed_count = 0
        
        for filename in individual_files:
            try:
                file_path = Path(filename)
                if file_path.exists():
                    file_path.unlink()  # Delete the file
                    logger.info(f"Removed individual file: {filename}")
                    removed_count += 1
                else:
                    logger.warning(f"File not found for cleanup: {filename}")
            except Exception as e:
                logger.error(f"Error removing file {filename}: {e}")
        
        logger.info(f"Cleanup completed. Removed {removed_count} individual race files.")
    
def main():
    # Configuration
    race_date = date(2026, 4, 29)
    race_course = "HV"  # ST / HV
    
    # Create downloader and process
    downloader = HKJCRacecardDownloader(race_date, race_course)
    downloader.download_all_racecards()
    
    logger.info("Racecard download completed!")

if __name__ == "__main__":
    main()