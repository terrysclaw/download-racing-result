import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import date
import numpy as np
import logging
from typing import Dict, List, Optional, Tuple
from pathlib import Path
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from horse_racing_analysis import RacingAnalyzer

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class HKJCRacecardDownloader:
    def __init__(self, race_date: date, race_course: str):
        self.race_date = race_date
        self.race_course = race_course
        self.base_url = "https://racing.hkjc.com/racing/information/Chinese/Racing"
        self.df_results = self._load_historical_data()
        # Initialize racing analyzer with historical data
        self.analyzer = RacingAnalyzer(self.df_results)
        
    def _parse_time_string(self, time_str: str) -> Optional[float]:
        """Unified time string parsing - converts time string to seconds.
        
        Args:
            time_str: Time string in format 'MM:SS.ms' or 'SS.ms' or float
            
        Returns:
            Float representing seconds, or None if parsing fails
        """
        if pd.isna(time_str) or time_str == '' or time_str == '---':
            return None
        
        try:
            time_str = str(time_str).strip()
            if ':' in time_str:
                # Format: MM:SS.ms
                parts = time_str.split(':')
                minutes = int(parts[0])
                seconds = float(parts[1])
                return minutes * 60 + seconds
            else:
                # Format: SS.ms or just float
                return float(time_str)
        except (ValueError, IndexError, AttributeError, TypeError):
            logger.debug(f"Failed to parse time string: {time_str}")
            return None
    
    def _get_race_count(self) -> int:
        """Dynamically detect the number of races for the configured date and course.
        
        Returns:
            Integer representing the total number of races, defaults to 11 if detection fails
        """
        try:
            url = f"https://racing.hkjc.com/racing/information/Chinese/Racing/RaceFixture.aspx?RaceDate={self.race_date.strftime('%Y/%m/%d')}&Racecourse={self.race_course}"
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Find the race number dropdown/select options
            race_selectors = soup.find_all('option')
            race_numbers = []
            
            for option in race_selectors:
                try:
                    # Try to extract race number from option value
                    value = option.get('value', '')
                    if value and value.isdigit():
                        race_numbers.append(int(value))
                except (ValueError, TypeError):
                    continue
            
            if race_numbers:
                max_race = max(race_numbers)
                logger.info(f"Detected {max_race} races for {self.race_date} at {self.race_course}")
                return max_race
            else:
                logger.warning(f"Could not detect race count, defaulting to 11")
                return 11
                
        except Exception as e:
            logger.warning(f"Error detecting race count: {e}, defaulting to 11")
            return 11
    
    def _load_historical_data(self) -> pd.DataFrame:
        """Load historical race results."""
        try:
            df_2024 = pd.read_excel('raceresult_2024.xlsx')
            df_2025 = pd.read_excel('raceresult_2025.xlsx')
            return pd.concat([df_2024, df_2025], ignore_index=True)
        except FileNotFoundError as e:
            logger.error(f"Historical data file not found: {e}")
            return pd.DataFrame()
    
    def _initialize_data_structure(self) -> Dict[str, List]:
        """Initialize the data structure for racecard."""
        return {
            '日期': [], '馬場': [], '場次': [], '泥草': [], '賽道': [], '班次': [], '路程': [],
            '馬匹編號': [], '6次近績': [], '馬名': [], '網址': [], '烙號': [], '負磅': [],
            '騎師': [], '可能超磅': [], '檔位': [], '練馬師': [], '國際評分': [], '評分': [],
            '評分+/-': [], '排位體重': [], '排位體重+/-': [], '最佳時間': [], '馬齡': [],
            '分齡讓磅': [], '性別': [], '今季獎金': [], '優先參賽次序': [], '上賽距今日數': [],
            '配備': [], '馬主': [], '父系': [], '母系': [], '進口類別': [],
        }
    
    def _safe_int_convert(self, value: str) -> Optional[int]:
        """Safely convert string to integer."""
        try:
            return int(value.strip())
        except (ValueError, AttributeError):
            return None
    
    def _clean_text(self, text: str) -> str:
        """Clean text by removing unwanted characters."""
        return text.replace('\r', '').replace('\n', '').replace('\xa0', '').strip()
    
    def _extract_race_info(self, soup: BeautifulSoup) -> Tuple[Optional[str], Optional[str], Optional[int]]:
        """Extract race information from soup."""
        race_div = soup.find('div', class_='f_fs13')
        
        if not race_div:
            return None, None, None
            
        band = None
        race_class = None
        distance = None
        
        for col in race_div.text.split(','):
            if '賽道' in col:
                band = col.strip()
            if '班' in col:
                race_class = col.strip()
            if '米' in col:
                try:
                    distance = self._safe_int_convert(col.split('米')[0])
                except:
                    distance = None
                    
        return band, race_class, distance
    
    def _extract_horse_data(self, cells: List, race_info: Dict, data: Dict[str, List]) -> None:
        """Extract individual horse data from table cells."""
        # Add race information
        for key, value in race_info.items():
            data[key].append(value)
        
        # Extract horse-specific data with safe conversion
        data['馬匹編號'].append(self._safe_int_convert(cells[0].text))
        data['6次近績'].append(self._clean_text(cells[1].text) if len(cells) > 1 else None)
        
        # Horse name and URL
        if len(cells) > 3:
            data['馬名'].append(self._clean_text(cells[3].text))
            try:
                data['網址'].append("https://racing.hkjc.com" + cells[3].find('a')['href'])
            except:
                data['網址'].append(None)
        else:
            data['馬名'].append(None)
            data['網址'].append(None)
        
        # Continue with other fields
        fields_mapping = [
            ('烙號', 4), ('負磅', 5), ('騎師', 6), ('可能超磅', 7), ('檔位', 8),
            ('練馬師', 9), ('國際評分', 10), ('評分', 11), ('評分+/-', 12),
            ('排位體重', 13), ('排位體重+/-', 14), ('最佳時間', 15), ('馬齡', 16),
            ('分齡讓磅', 17), ('性別', 18), ('今季獎金', 19), ('優先參賽次序', 20),
            ('上賽距今日數', 21), ('配備', 22), ('馬主', 23), ('父系', 24),
            ('母系', 25), ('進口類別', 26)
        ]
        
        for field_name, index in fields_mapping:
            if index < len(cells):
                if field_name in ['負磅', '檔位', '國際評分', '評分', '排位體重', '馬齡', '分齡讓磅']:
                    data[field_name].append(self._safe_int_convert(cells[index].text))
                else:
                    data[field_name].append(self._clean_text(cells[index].text))
            else:
                data[field_name].append(None)
    
    def _get_time_adjustment_factors(self, venue: str, surface: str, distance: int) -> Dict[str, float]:
        """Get time adjustment factors based on venue, surface, and distance."""
        factors = {'weight_early': 0, 'weight_late': 0, 'gate': 0}
        
        if venue == 'HV':
            distance_factors = {
                1000: {'weight_early': 0.015, 'weight_late': 0.056, 'gate': 0.02},
                1200: {'weight_early': 0.017, 'weight_late': 0.056, 'gate': 0.033},
                1650: {'weight_early': 0.019, 'weight_late': 0.056, 'gate': 0.043},
                1800: {'weight_early': 0.021, 'weight_late': 0.056, 'gate': 0.048},
                2200: {'weight_early': 0.023, 'weight_late': 0.056, 'gate': 0.053}
            }
        else:  # ST
            if surface == '草地':
                distance_factors = {
                    1000: {'weight_early': 0.015, 'weight_late': 0.056, 'gate': -0.019},
                    1200: {'weight_early': 0.017, 'weight_late': 0.056, 'gate': 0.028},
                    1400: {'weight_early': 0.018, 'weight_late': 0.056, 'gate': 0.033},
                    1600: {'weight_early': 0.019, 'weight_late': 0.056, 'gate': 0.038},
                    1800: {'weight_early': 0.021, 'weight_late': 0.056, 'gate': 0.043},
                    2000: {'weight_early': 0.022, 'weight_late': 0.056, 'gate': 0.048},
                    2400: {'weight_early': 0.024, 'weight_late': 0.056, 'gate': 0.053}
                }
            else:  # 泥地
                distance_factors = {
                    1200: {'weight_early': 0.017, 'weight_late': 0.056, 'gate': 0.028},
                    1650: {'weight_early': 0.019, 'weight_late': 0.056, 'gate': 0.043},
                    1800: {'weight_early': 0.021, 'weight_late': 0.056, 'gate': 0.043}
                }
        
        return distance_factors.get(distance, factors)
    
    def _calculate_time_adjustment(self, row: pd.Series) -> float:
        """Calculate time adjustment based on weight and gate changes."""
        if pd.isna(row['上次負磅 +/-']) or pd.isna(row['上次檔位 +/-']):
            return 0
        
        factors = self._get_time_adjustment_factors(
            row['馬場'], row['泥草'], row['路程']
        )
        
        distance = row['路程']
        weight_diff = row['上次負磅 +/-']
        gate_diff = row['上次檔位 +/-']
        
        # Calculate weight adjustment (early and late portions)
        early_portion = distance - 400 if distance > 400 else 0
        late_portion = min(distance, 400)
        
        time_adjustment = 0
        time_adjustment += (weight_diff * factors['weight_early'] / distance * early_portion)
        time_adjustment += (weight_diff * factors['weight_late'] / distance * late_portion)
        time_adjustment += gate_diff * factors['gate']
        
        return time_adjustment
    
    def _apply_ground_adjustment(self, time_value: float, ground_condition: str, venue: str) -> float:
        """Apply ground condition adjustment to time."""
        if ground_condition == '好地至快地':
            multiplier = 1.006 if venue == 'HV' else 1.004
            return time_value * multiplier
        elif ground_condition == '好地至黏地':
            return time_value / 1.012
        return time_value
    
    def _calculate_sectional_adjustments(self, row: pd.Series) -> Optional[float]:
        """Calculate adjusted last 800m time."""
        if pd.isna(row['上次最後 800']) or pd.isna(row['上次負磅 +/-']) or pd.isna(row['上次檔位 +/-']):
            return None
        
        time_1 = row['上次彎位 400'] if pd.notna(row['上次彎位 400']) else 0
        time_2 = row['上次最後 400'] if pd.notna(row['上次最後 400']) else 0
        
        factors = self._get_time_adjustment_factors(
            row['馬場'], row['泥草'], row['路程']
        )
        
        distance = row['路程']
        weight_diff = row['上次負磅 +/-']
        gate_diff = row['上次檔位 +/-']
        
        # Adjust each 400m section
        time_1 += (weight_diff * factors['weight_early'] / distance * 400)
        time_1 += gate_diff * factors['gate'] / distance * 400
        time_2 += (weight_diff * factors['weight_late'] / distance * 400)
        
        adjusted_800 = time_1 + time_2
        
        # Apply ground condition adjustment
        if row['上次泥草'] == '草地':
            adjusted_800 = self._apply_ground_adjustment(
                adjusted_800, row['上次場地狀況'], row['上次馬場']
            )
        
        return round(adjusted_800, 3)
    
    def _process_historical_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Process and add historical performance data."""
        # Initialize columns
        historical_columns = [
            '上次總場次', '上次日期', '上次馬場', '上次泥草', '上次場次', '上次班次', '上次路程',
            '上次場地狀況', '上次名次', '上次騎師', '上次賠率', '上次負磅', '上次負磅 +/-',
            '上次檔位', '上次檔位 +/-', '上次賽事時間1', '上次賽事時間2', '上次賽事時間3',
            '上次賽事時間4', '上次賽事時間5', '上次賽事時間6', '上次賽事時間', '上次完成時間',
            '上次第 1 段', '上次第 2 段', '上次第 3 段', '上次第 4 段', '上次第 5 段', '上次第 6 段',
            '上次最後 400', '上次最後 800', '上次頭段完成時間', '上次頭段', '上次彎位 400',
            '上次調整基數', '上次調整後最後 800', '上次調整後完成時間', '上次調整後完成秒速',

            '前次總場次', '前次日期', '前次馬場', '前次泥草', '前次場次', '前次班次', '前次路程',
            '前次場地狀況', '前次名次', '前次騎師', '前次賠率', '前次負磅', '前次負磅 +/-',
            '前次檔位', '前次檔位 +/-', '前次賽事時間1', '前次賽事時間2', '前次賽事時間3',
            '前次賽事時間4', '前次賽事時間5', '前次賽事時間6', '前次賽事時間', '前次完成時間',
            '前次第 1 段', '前次第 2 段', '前次第 3 段', '前次第 4 段', '前次第 5 段', '前次第 6 段',
            '前次最後 400', '前次最後 800', '前次頭段完成時間', '前次頭段', '前次彎位 400',
            '前次調整基數', '前次調整後最後 800', '前次調整後完成時間', '前次調整後完成秒速'
        ]
        
        for col in historical_columns:
            df[col] = None
        
        for index, row in df.iterrows():
            # Find last matching race
            if row['泥草'] == '草地':
                last_match = self.df_results[
                    (self.df_results['布號'] == row['烙號']) &
                    (self.df_results['馬場'] == row['馬場']) &
                    (self.df_results['泥草'] == row['泥草']) &
                    (~self.df_results['場地狀況'].isin(['軟地', '黏地'])) &
                    (self.df_results['路程'] == row['路程']) &
                    (self.df_results['獨贏賠率'] > 0)
                ].tail(2)
            else:
                last_match = self.df_results[
                    (self.df_results['布號'] == row['烙號']) &
                    (self.df_results['馬場'] == row['馬場']) &
                    (self.df_results['泥草'] == row['泥草']) &
                    (self.df_results['路程'] == row['路程']) &
                    (self.df_results['獨贏賠率'] > 0)
                ].tail(2)
            
            if not last_match.empty:
                # self._populate_historical_data(df, index, row, last_match.iloc[0])
                self._populate_historical_data(df, index, row, last_match)
        
        return df
    
    def _populate_historical_data(self, df: pd.DataFrame, index: int, row: pd.Series, last_match: pd.DataFrame) -> None:
        """Populate historical data for a horse."""
        # Reverse to process most recent first
        for i, (_, match) in enumerate(last_match.iloc[::-1].iterrows(), start=1):
            prefix = '上次' if i == 1 else '前次'
            self._populate_single_historical_data(df, index, row, match, prefix)

    def _populate_single_historical_data(self, df: pd.DataFrame, index: int, row: pd.Series, last_match: pd.Series, prefix: str) -> None:
        # Basic historical data
        df.at[index, f'{prefix}總場次'] = last_match['總場次']
        df.at[index, f'{prefix}日期'] = last_match['日期']
        df.at[index, f'{prefix}馬場'] = last_match['馬場']
        df.at[index, f'{prefix}泥草'] = last_match['泥草']
        df.at[index, f'{prefix}場次'] = last_match['場次']
        df.at[index, f'{prefix}班次'] = last_match['班次']
        df.at[index, f'{prefix}路程'] = last_match['路程']
        df.at[index, f'{prefix}場地狀況'] = last_match['場地狀況']
        df.at[index, f'{prefix}名次'] = last_match['名次']
        df.at[index, f'{prefix}騎師'] = last_match['騎師']
        df.at[index, f'{prefix}賠率'] = last_match['獨贏賠率']
        df.at[index, f'{prefix}負磅'] = last_match['實際負磅']
        df.at[index, f'{prefix}檔位'] = last_match['檔位']
        
        # Calculate differences
        if pd.notna(last_match['實際負磅']):
            df.at[index, f'{prefix}負磅 +/-'] = row['負磅'] - last_match['實際負磅']
        
        if pd.notna(last_match['檔位']):
            df.at[index, f'{prefix}檔位 +/-'] = row['檔位'] - last_match['檔位']
        
        # Race times
        for i in range(1, 7):
            df.at[index, f'{prefix}賽事時間{i}'] = last_match[f'賽事時間{i}']
            df.at[index, f'{prefix}第 {i} 段'] = last_match[f'第 {i} 段']
        
        # Set appropriate race time based on distance
        distance = row['路程']
        if 1000 <= distance <= 1200:
            df.at[index, f'{prefix}賽事時間'] = last_match['賽事時間3']
        elif 1200 < distance <= 1650:
            df.at[index, f'{prefix}賽事時間'] = last_match['賽事時間4']
        elif 1650 < distance <= 2000:
            df.at[index, f'{prefix}賽事時間'] = last_match['賽事時間5']
        elif distance > 2000:
            df.at[index, f'{prefix}賽事時間'] = last_match['賽事時間6']
        
        df.at[index, f'{prefix}完成時間'] = last_match['完成時間']
        
        # Calculate sectional times
        self._calculate_sectional_times(df, index, last_match, prefix)
        
        # Calculate time adjustments
        if pd.notna(df.at[index, f'{prefix}完成時間']) and df.at[index, f'{prefix}完成時間'] != '---':
            self._calculate_adjusted_times(df, index, row, prefix)
    
    def _calculate_sectional_times(self, df: pd.DataFrame, index: int, last_match: pd.Series, prefix: str) -> None:
        """Calculate sectional time breakdowns."""
        # Find the last available sectional segment
        available_sections = []
        for i in range(1, 7):
            section_value = last_match.get(f'第 {i} 段')
            if pd.notna(section_value) and section_value != '' and section_value != '---':
                try:
                    # Ensure it's a valid numeric value
                    float(section_value)
                    available_sections.append(i)
                except (ValueError, TypeError):
                    pass
    
        if not available_sections:
            return
        
        last_section = max(available_sections)
        
        def safe_float(value, default=0.0):
            """Safely convert value to float."""
            if pd.isna(value) or value == '' or value == '---':
                return default
            try:
                return float(value)
            except (ValueError, TypeError):
                return default
        
        def safe_sum(values):
            """Safely sum a list of values."""
            return round(sum(safe_float(v) for v in values), 2)
        
        # Calculate sectional times based on available data
        if last_section >= 6:  # 6 sections available
            df.at[index, f'{prefix}頭段完成時間'] = last_match.get('賽事時間5')
            df.at[index, f'{prefix}頭段'] = safe_sum([last_match.get(f'第 {j} 段') for j in range(1, 5)])
            df.at[index, f'{prefix}彎位 400'] = round(safe_float(last_match.get('第 5 段')), 2)
            df.at[index, f'{prefix}最後 400'] = round(safe_float(last_match.get('第 6 段')), 2)
            df.at[index, f'{prefix}最後 800'] = round(safe_float(last_match.get('第 5 段')) + safe_float(last_match.get('第 6 段')), 2)
            
        elif last_section >= 5:  # 5 sections available
            df.at[index, f'{prefix}頭段完成時間'] = last_match.get('賽事時間4')
            df.at[index, f'{prefix}頭段'] = safe_sum([last_match.get(f'第 {j} 段') for j in range(1, 4)])
            df.at[index, f'{prefix}彎位 400'] = round(safe_float(last_match.get('第 4 段')), 2)
            df.at[index, f'{prefix}最後 400'] = round(safe_float(last_match.get('第 5 段')), 2)
            df.at[index, f'{prefix}最後 800'] = round(safe_float(last_match.get('第 4 段')) + safe_float(last_match.get('第 5 段')), 2)
            
        elif last_section >= 4:  # 4 sections available
            df.at[index, f'{prefix}頭段完成時間'] = last_match.get('賽事時間3')
            df.at[index, f'{prefix}頭段'] = safe_sum([last_match.get(f'第 {j} 段') for j in range(1, 3)])
            df.at[index, f'{prefix}彎位 400'] = round(safe_float(last_match.get('第 3 段')), 2)
            df.at[index, f'{prefix}最後 400'] = round(safe_float(last_match.get('第 4 段')), 2)
            df.at[index, f'{prefix}最後 800'] = round(safe_float(last_match.get('第 3 段')) + safe_float(last_match.get('第 4 段')), 2)
            
        elif last_section >= 3:  # 3 sections available
            df.at[index, f'{prefix}頭段完成時間'] = last_match.get('賽事時間2')
            df.at[index, f'{prefix}頭段'] = safe_sum([last_match.get('第 1 段'), last_match.get('第 2 段')])
            df.at[index, f'{prefix}彎位 400'] = round(safe_float(last_match.get('第 2 段')), 2)
            df.at[index, f'{prefix}最後 400'] = round(safe_float(last_match.get('第 3 段')), 2)
            df.at[index, f'{prefix}最後 800'] = round(safe_float(last_match.get('第 2 段')) + safe_float(last_match.get('第 3 段')), 2)
            
        elif last_section >= 2:  # Only 2 sections available
            df.at[index, f'{prefix}頭段完成時間'] = last_match.get('賽事時間1')
            df.at[index, f'{prefix}頭段'] = round(safe_float(last_match.get('第 1 段')), 2)
            df.at[index, f'{prefix}彎位 400'] = round(safe_float(last_match.get('第 1 段')), 2)
            df.at[index, f'{prefix}最後 400'] = round(safe_float(last_match.get('第 2 段')), 2)
            df.at[index, f'{prefix}最後 800'] = round(safe_float(last_match.get('第 1 段')) + safe_float(last_match.get('第 2 段')), 2)
            
        else:  # Only 1 section available
            df.at[index, f'{prefix}頭段完成時間'] = None
            df.at[index, f'{prefix}頭段'] = None
            df.at[index, f'{prefix}彎位 400'] = None
            df.at[index, f'{prefix}最後 400'] = round(safe_float(last_match.get('第 1 段')), 2)
            df.at[index, f'{prefix}最後 800'] = None
        
        # Set adjustment base (調整基數)
        try:
            df.at[index, f'{prefix}調整基數'] = self._calculate_time_adjustment(df.loc[index])
        except Exception:
            df.at[index, f'{prefix}調整基數'] = 0
    
    def _calculate_adjusted_times(self, df: pd.DataFrame, index: int, row: pd.Series, prefix: str) -> None:
        """Calculate time-adjusted performance metrics."""
        finish_time_str = df.at[index, f'{prefix}完成時間']
        
        try:
            # Convert finish time to seconds
            time_parts = finish_time_str.split(':')
            finish_seconds = int(time_parts[0]) * 60 + float(time_parts[1])
            
            # Apply time adjustment
            time_adjustment = self._calculate_time_adjustment(df.loc[index])
            adjusted_seconds = finish_seconds + time_adjustment
            
            # Apply ground condition adjustment
            if df.at[index, f'{prefix}泥草'] == '草地':
                adjusted_seconds = self._apply_ground_adjustment(
                    adjusted_seconds, df.at[index, f'{prefix}場地狀況'], df.at[index, f'{prefix}馬場']
                )
            
            # Store results
            df.at[index, f'{prefix}調整後完成秒速'] = adjusted_seconds
            df.at[index, f'{prefix}調整後完成時間'] = f"{int(adjusted_seconds // 60)}:{round(adjusted_seconds % 60, 2)}"
            
            # Calculate adjusted last 800m
            df.at[index, f'{prefix}調整後最後 800'] = self._calculate_sectional_adjustments(df.loc[index])
            
        except (ValueError, IndexError) as e:
            logger.warning(f"Error processing finish time for index {index}: {e}")
    
    def _scrape_race_data(self, race_no: int) -> Optional[pd.DataFrame]:
        """Scrape race data for a specific race."""
        url = f"{self.base_url}/RaceCard.aspx?RaceDate={self.race_date.strftime('%Y/%m/%d')}&Racecourse={self.race_course}&RaceNo={race_no}"
        
        try:
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Extract race information
            band, race_class, distance = self._extract_race_info(soup)
            
            race_info = {
                '日期': self.race_date,
                '馬場': self.race_course,
                '場次': race_no,
                '泥草': '草地' if band else '泥地',
                '賽道': band,
                '班次': race_class,
                '路程': distance
            }
            
            # Find racecard table
            racecard_table = soup.find('table', id='racecardlist')
            if not racecard_table:
                logger.warning(f"No racecard table found for Race {race_no}")
                return None
            
            # Extract horse data
            data = self._initialize_data_structure()
            tbody = racecard_table.find('tbody')
            
            if tbody:
                for tr in tbody.find_all('tr')[2:]:  # Skip header rows
                    cells = tr.find_all('td')
                    if len(cells) >= 4:  # Ensure minimum required cells
                        self._extract_horse_data(cells, race_info, data)
            
            df = pd.DataFrame(data)
            
            if not df.empty:
                df = self._process_historical_data(df)
            
            return df
            
        except requests.RequestException as e:
            logger.error(f"Error fetching race {race_no}: {e}")
            return None
        except Exception as e:
            logger.error(f"Unexpected error processing race {race_no}: {e}")
            return None
    
    def _create_terry_sheet(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create Terry AI formatted sheet."""
        # First, ensure all required columns exist in the DataFrame
        required_columns = [
            '場次', '班次', '路程', '馬匹編號', '馬名', '騎師', '練馬師', '評分', '6次近績',
            '烙號', '泥草', '負磅', '排位體重', '排位體重+/-',
            '上次日期', '上賽距今日數', '上次場次', '上次班次', '上次場地狀況',
            '上次負磅', '上次負磅 +/-', '檔位', '上次檔位', '上次檔位 +/-',
            '最佳時間', '上次賽事時間', '上次完成時間',
            '上次調整後完成時間', '上次調整後完成秒速', '上次頭段完成時間', '上次頭段',
            '上次彎位 400', '上次最後 400', '上次最後 800', '上次調整後最後 800',
            '前次日期', '前次場次', '前次班次', '前次場地狀況',
            '前次負磅', '前次負磅 +/-', '前次檔位', '前次檔位 +/-',
            '前次賽事時間', '前次完成時間',
            '前次調整後完成時間', '前次調整後完成秒速', '前次頭段完成時間', '前次頭段',
            '前次彎位 400', '前次最後 400', '前次最後 800', '前次調整後最後 800',
            '網址'
        ]
        
        # Add missing columns that will be calculated
        calculated_columns = [
            '上次完成時間 - 上次賽事時間', '賽事結果連結', '賽事重溫連結',
            '前次完成時間 - 前次賽事時間', '前次賽事結果連結', '前次賽事重溫連結',
            '勝率(%)', '騎師因子', '一致性評分', '綜合評分', '評估',
            '騎練合作往績因子', '體重變化(磅)', '體重狀態警示', '負磅百分比(%)',
            '步速預測', '近3仗頭段均速', '特徵修正勝率(%)',
            '較快調整後完成秒速', '較快調整後最後 800', '預測得分', '預測排名'
        ]
        
        # Create a copy of the DataFrame with only existing columns
        existing_columns = [col for col in required_columns if col in df.columns]
        df_terry = df[existing_columns].copy()
        # Convert all columns to object dtype to allow flexible type assignment
        df_terry = df_terry.astype('object')
        
        # Add calculated columns
        for col in calculated_columns:
            df_terry[col] = None
        
        race_date_ts = pd.Timestamp(self.race_date)
        
        for index, row in df_terry.iterrows():
            # ========== 計算勝率、騎師因子等投注指標 ==========
            # Get field size (count of horses in the race)
            field_size = len([r for r in df_terry['馬匹編號'] if pd.notna(r)])
            field_size = max(field_size, 3)  # Minimum 3 horses

            rating = row.get('評分')
            jockey = row.get('騎師')
            trainer = row.get('練馬師')
            form_6 = row.get('6次近績')

            # Feature 1: Trainer-Jockey combo factor
            combo_factor = 1.0
            if pd.notna(trainer) and trainer != '' and pd.notna(jockey) and jockey != '':
                try:
                    combo_factor = self.analyzer.get_trainer_jockey_combo_factor(str(trainer), str(jockey))
                    df_terry.at[index, '騎練合作往績因子'] = round(combo_factor, 2)
                except Exception as e:
                    logger.debug(f"Error getting trainer-jockey combo factor: {e}")

            # Feature 2: Body weight change warning
            weight_change = row.get('排位體重+/-')
            if pd.notna(weight_change):
                try:
                    weight_change_val = float(weight_change)
                    df_terry.at[index, '體重變化(磅)'] = round(weight_change_val, 1)
                    if abs(weight_change_val) >= 20:
                        df_terry.at[index, '體重狀態警示'] = 'WARNING (>=20lb)'
                    else:
                        df_terry.at[index, '體重狀態警示'] = 'OK'
                except (TypeError, ValueError):
                    pass

            # Feature 3: Load ratio (assigned weight / body weight)
            assigned_weight = row.get('負磅')
            body_weight = row.get('排位體重')
            load_ratio_pct = None
            if pd.notna(assigned_weight) and pd.notna(body_weight):
                try:
                    body_weight_val = float(body_weight)
                    if body_weight_val > 0:
                        load_ratio_pct = float(assigned_weight) / body_weight_val * 100.0
                        df_terry.at[index, '負磅百分比(%)'] = round(load_ratio_pct, 2)
                except (TypeError, ValueError, ZeroDivisionError):
                    pass

            # Feature 4: Pace mapping from recent head-section times
            pace_times = []
            for t_col in ['上次頭段完成時間', '前次頭段完成時間']:
                parsed_time = self._parse_time_string(row.get(t_col))
                if parsed_time is not None:
                    pace_times.append(parsed_time)

            cloth_no = row.get('烙號')
            if len(pace_times) < 3 and pd.notna(cloth_no):
                try:
                    historical_pace = self.analyzer.get_recent_head_section_times(
                        cloth_no=str(cloth_no),
                        venue=self.race_course,
                        surface=str(row.get('泥草')) if pd.notna(row.get('泥草')) else None,
                        distance=int(row.get('路程')) if pd.notna(row.get('路程')) else None,
                        limit=3
                    )
                    for p in historical_pace:
                        if len(pace_times) >= 3:
                            break
                        pace_times.append(float(p))
                except Exception as e:
                    logger.debug(f"Error building pace mapping data: {e}")

            if pace_times:
                pace_style = self.analyzer.classify_pace_style(pace_times)
                df_terry.at[index, '步速預測'] = pace_style
                df_terry.at[index, '近3仗頭段均速'] = round(float(np.mean(pace_times[:3])), 2)
            
            # Calculate base win probability
            if pd.notna(rating) and isinstance(rating, (int, float)):
                try:
                    win_prob = self.analyzer.calculate_win_probability(
                        int(rating), 
                        str(row.get('班次', '')), 
                        field_size
                    )
                    df_terry.at[index, '勝率(%)'] = round(win_prob * 100, 1)
                except Exception as e:
                    logger.debug(f"Error calculating win probability: {e}")
            
            # Get jockey factor
            if pd.notna(jockey) and jockey != '':
                try:
                    jockey_factor = self.analyzer.get_jockey_factor(str(jockey))
                    df_terry.at[index, '騎師因子'] = round(jockey_factor, 2)
                except Exception as e:
                    logger.debug(f"Error getting jockey factor: {e}")

            # Calculate feature-adjusted win probability
            if pd.notna(rating) and isinstance(rating, (int, float)):
                try:
                    base_prob = self.analyzer.calculate_win_probability(
                        int(rating),
                        str(row.get('班次', '')),
                        field_size
                    )
                    pace_style = df_terry.at[index, '步速預測']
                    pace_multiplier_map = {
                        'Front-runner': 1.08,
                        'Prominent': 1.03,
                        'Closer': 0.97,
                        'Unknown': 1.00
                    }
                    pace_multiplier = pace_multiplier_map.get(pace_style, 1.00)

                    weight_multiplier = 1.00
                    if pd.notna(weight_change):
                        try:
                            if abs(float(weight_change)) >= 20:
                                weight_multiplier = 0.88
                        except (TypeError, ValueError):
                            pass

                    load_multiplier = 1.00
                    if load_ratio_pct is not None:
                        if load_ratio_pct >= 11.5:
                            load_multiplier = 0.95
                        elif load_ratio_pct <= 10.0:
                            load_multiplier = 1.03

                    adjusted_prob = base_prob * combo_factor * pace_multiplier * weight_multiplier * load_multiplier
                    adjusted_prob = max(0.0, min(0.95, adjusted_prob))
                    df_terry.at[index, '特徵修正勝率(%)'] = round(adjusted_prob * 100.0, 1)
                except Exception as e:
                    logger.debug(f"Error calculating feature-adjusted win probability: {e}")
            
            # Calculate consistency score from recent form
            if pd.notna(form_6) and form_6 != '':
                try:
                    consistency_score, _ = self.analyzer.calculate_consistency_score(str(form_6))
                    df_terry.at[index, '一致性評分'] = round(consistency_score, 1)
                except Exception as e:
                    logger.debug(f"Error calculating consistency score: {e}")
            
            # Calculate composite score
            if pd.notna(rating) and pd.notna(jockey) and pd.notna(form_6):
                try:
                    composite = self.analyzer.calculate_composite_score(
                        int(rating),
                        str(jockey),
                        str(form_6),
                        str(row.get('班次', '')),
                        field_size
                    )
                    if composite:
                        df_terry.at[index, '綜合評分'] = round(composite.get('綜合評分', 0), 1)
                        df_terry.at[index, '評估'] = composite.get('評估', '')
                except Exception as e:
                    logger.debug(f"Error calculating composite score: {e}")
            
            # ========== 原有的時間計算邏輯 ==========
            # Process Last Race (上次)
            if pd.notna(row.get('上次日期')):
                last_date = pd.to_datetime(row['上次日期'])
                df_terry.at[index, '上賽距今日數'] = (race_date_ts - last_date).days
                
                # Calculate time difference
                if pd.notna(row.get('上次完成時間')) and pd.notna(row.get('上次賽事時間')):
                    finish_time = self._parse_time_string(row['上次完成時間'])
                    race_time = self._parse_time_string(row['上次賽事時間'])
                    
                    if finish_time is not None and race_time is not None:
                        df_terry.at[index, '上次完成時間 - 上次賽事時間'] = finish_time - race_time
                    else:
                        df_terry.at[index, '上次完成時間 - 上次賽事時間'] = None
                
                # Add result and replay links
                if pd.notna(row.get('上次場次')):
                    try:
                        race_no = int(row['上次場次'])
                        df_terry.at[index, '賽事結果連結'] = f"https://racing.hkjc.com/racing/information/chinese/Racing/LocalResults.aspx?RaceDate={last_date.strftime('%Y/%m/%d')}&RaceNo={race_no}"
                        df_terry.at[index, '賽事重溫連結'] = f"https://racing.hkjc.com/racing/video/play.asp?type=replay-full&date={last_date.strftime('%Y%m%d')}&no={race_no:02d}&lang=chi"
                    except (ValueError, TypeError):
                        pass

            # Process Previous Race (前次)
            if pd.notna(row.get('前次日期')):
                prev_date = pd.to_datetime(row['前次日期'])
                
                # Calculate time difference
                if pd.notna(row.get('前次完成時間')) and pd.notna(row.get('前次賽事時間')):
                    finish_time = self._parse_time_string(row['前次完成時間'])
                    race_time = self._parse_time_string(row['前次賽事時間'])
                    
                    if finish_time is not None and race_time is not None:
                        df_terry.at[index, '前次完成時間 - 前次賽事時間'] = finish_time - race_time
                    else:
                        df_terry.at[index, '前次完成時間 - 前次賽事時間'] = None
                
                # Add result and replay links
                if pd.notna(row.get('前次場次')):
                    try:
                        race_no = int(row['前次場次'])
                        df_terry.at[index, '前次賽事結果連結'] = f"https://racing.hkjc.com/racing/information/chinese/Racing/LocalResults.aspx?RaceDate={prev_date.strftime('%Y/%m/%d')}&RaceNo={race_no}"
                        df_terry.at[index, '前次賽事重溫連結'] = f"https://racing.hkjc.com/racing/video/play.asp?type=replay-full&date={prev_date.strftime('%Y%m%d')}&no={race_no:02d}&lang=chi"
                    except (ValueError, TypeError):
                        pass
            
            # Process 兩次時間那一次較快
            last_adjusted = row.get('上次調整後完成秒速')
            prev_adjusted = row.get('前次調整後完成秒速')
            if pd.notna(last_adjusted) and pd.notna(prev_adjusted):
                df_terry.at[index, '較快調整後完成秒速'] = min(last_adjusted, prev_adjusted)
            elif pd.notna(last_adjusted):
                df_terry.at[index, '較快調整後完成秒速'] = last_adjusted
            elif pd.notna(prev_adjusted):
                df_terry.at[index, '較快調整後完成秒速'] = prev_adjusted

            # Process 兩次最後800時間那一次較快
            last_800 = row.get('上次調整後最後 800')
            prev_800 = row.get('前次調整後最後 800')
            if pd.notna(last_800) and pd.notna(prev_800):
                df_terry.at[index, '較快調整後最後 800'] = min(last_800, prev_800)
            elif pd.notna(last_800):
                df_terry.at[index, '較快調整後最後 800'] = last_800
            elif pd.notna(prev_800):
                df_terry.at[index, '較快調整後最後 800'] = prev_800

        return df_terry
    
    def _create_alfred_sheet(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create Alfred AI formatted sheet."""
        required_columns = [
            '場次', '班次', '路程', '馬匹編號', '馬名', '騎師', '練馬師', '評分',
            '烙號', '泥草', '負磅', '排位體重', '排位體重+/-',
            '上次日期', '上賽距今日數', '上次班次', '上次場地狀況',
            '上次負磅', '上次負磅 +/-', '檔位', '上次檔位', '上次檔位 +/-',
            '最佳時間', '上次完成時間', '上次頭段完成時間', '上次彎位 400',
            '上次最後 400', '上次最後 800', '前次頭段完成時間', '網址', '上次場次'
        ]
        
        # Add missing columns that will be calculated
        calculated_columns = [
            '賽事結果連結', '賽事重溫連結', '勝率(%)', '騎師因子',
            '騎練合作往績因子', '體重變化(磅)', '體重狀態警示', '負磅百分比(%)',
            '步速預測', '近3仗頭段均速', '特徵修正勝率(%)', '預測得分', '預測排名'
        ]
        
        # Create a copy of the DataFrame with only existing columns
        existing_columns = [col for col in required_columns if col in df.columns]
        df_alfred = df[existing_columns].copy()
        # Convert all columns to object dtype to allow flexible type assignment
        df_alfred = df_alfred.astype('object')
        
        # Add calculated columns
        for col in calculated_columns:
            df_alfred[col] = None
        
        race_date_ts = pd.Timestamp(self.race_date)
        
        for index, row in df_alfred.iterrows():
            # ========== 計算勝率、騎師因子等投注指標 ==========
            field_size = len([r for r in df_alfred['馬匹編號'] if pd.notna(r)])
            field_size = max(field_size, 3)

            rating = row.get('評分')
            jockey = row.get('騎師')
            trainer = row.get('練馬師')

            combo_factor = 1.0
            if pd.notna(trainer) and trainer != '' and pd.notna(jockey) and jockey != '':
                try:
                    combo_factor = self.analyzer.get_trainer_jockey_combo_factor(str(trainer), str(jockey))
                    df_alfred.at[index, '騎練合作往績因子'] = round(combo_factor, 2)
                except Exception as e:
                    logger.debug(f"Error getting trainer-jockey combo factor: {e}")

            weight_change = row.get('排位體重+/-')
            if pd.notna(weight_change):
                try:
                    weight_change_val = float(weight_change)
                    df_alfred.at[index, '體重變化(磅)'] = round(weight_change_val, 1)
                    df_alfred.at[index, '體重狀態警示'] = 'WARNING (>=20lb)' if abs(weight_change_val) >= 20 else 'OK'
                except (TypeError, ValueError):
                    pass

            assigned_weight = row.get('負磅')
            body_weight = row.get('排位體重')
            load_ratio_pct = None
            if pd.notna(assigned_weight) and pd.notna(body_weight):
                try:
                    body_weight_val = float(body_weight)
                    if body_weight_val > 0:
                        load_ratio_pct = float(assigned_weight) / body_weight_val * 100.0
                        df_alfred.at[index, '負磅百分比(%)'] = round(load_ratio_pct, 2)
                except (TypeError, ValueError, ZeroDivisionError):
                    pass

            pace_times = []
            for t_col in ['上次頭段完成時間', '前次頭段完成時間']:
                parsed_time = self._parse_time_string(row.get(t_col))
                if parsed_time is not None:
                    pace_times.append(parsed_time)

            cloth_no = row.get('烙號')
            if len(pace_times) < 3 and pd.notna(cloth_no):
                try:
                    historical_pace = self.analyzer.get_recent_head_section_times(
                        cloth_no=str(cloth_no),
                        venue=self.race_course,
                        surface=str(row.get('泥草')) if pd.notna(row.get('泥草')) else None,
                        distance=int(row.get('路程')) if pd.notna(row.get('路程')) else None,
                        limit=3
                    )
                    for p in historical_pace:
                        if len(pace_times) >= 3:
                            break
                        pace_times.append(float(p))
                except Exception as e:
                    logger.debug(f"Error building pace mapping data: {e}")

            if pace_times:
                pace_style = self.analyzer.classify_pace_style(pace_times)
                df_alfred.at[index, '步速預測'] = pace_style
                df_alfred.at[index, '近3仗頭段均速'] = round(float(np.mean(pace_times[:3])), 2)
            
            # Calculate win probability
            if pd.notna(rating) and isinstance(rating, (int, float)):
                try:
                    win_prob = self.analyzer.calculate_win_probability(
                        int(rating), 
                        str(row.get('班次', '')), 
                        field_size
                    )
                    df_alfred.at[index, '勝率(%)'] = round(win_prob * 100, 1)
                except Exception as e:
                    logger.debug(f"Error calculating win probability: {e}")
            
            # Get jockey factor
            if pd.notna(jockey) and jockey != '':
                try:
                    jockey_factor = self.analyzer.get_jockey_factor(str(jockey))
                    df_alfred.at[index, '騎師因子'] = round(jockey_factor, 2)
                except Exception as e:
                    logger.debug(f"Error getting jockey factor: {e}")

            if pd.notna(rating) and isinstance(rating, (int, float)):
                try:
                    base_prob = self.analyzer.calculate_win_probability(
                        int(rating),
                        str(row.get('班次', '')),
                        field_size
                    )
                    pace_style = df_alfred.at[index, '步速預測']
                    pace_multiplier_map = {
                        'Front-runner': 1.08,
                        'Prominent': 1.03,
                        'Closer': 0.97,
                        'Unknown': 1.00
                    }
                    pace_multiplier = pace_multiplier_map.get(pace_style, 1.00)

                    weight_multiplier = 1.00
                    if pd.notna(weight_change):
                        try:
                            if abs(float(weight_change)) >= 20:
                                weight_multiplier = 0.88
                        except (TypeError, ValueError):
                            pass

                    load_multiplier = 1.00
                    if load_ratio_pct is not None:
                        if load_ratio_pct >= 11.5:
                            load_multiplier = 0.95
                        elif load_ratio_pct <= 10.0:
                            load_multiplier = 1.03

                    adjusted_prob = base_prob * combo_factor * pace_multiplier * weight_multiplier * load_multiplier
                    adjusted_prob = max(0.0, min(0.95, adjusted_prob))
                    df_alfred.at[index, '特徵修正勝率(%)'] = round(adjusted_prob * 100.0, 1)
                except Exception as e:
                    logger.debug(f"Error calculating feature-adjusted win probability: {e}")
            
            # ========== 原有的邏輯 ==========
            if pd.notna(row.get('上次日期')):
                last_date = pd.to_datetime(row['上次日期'])
                df_alfred.at[index, '上賽距今日數'] = (race_date_ts - last_date).days
                
                # Add result and replay links
                if pd.notna(row.get('上次場次')):
                    try:
                        race_no = int(row['上次場次'])
                        df_alfred.at[index, '賽事結果連結'] = f"https://racing.hkjc.com/racing/information/chinese/Racing/LocalResults.aspx?RaceDate={last_date.strftime('%Y/%m/%d')}&RaceNo={race_no}"
                        df_alfred.at[index, '賽事重溫連結'] = f"https://racing.hkjc.com/racing/video/play.asp?type=replay-full&date={last_date.strftime('%Y%m%d')}&no={race_no:02d}&lang=chi"
                    except (ValueError, TypeError):
                        pass
        
        return df_alfred

    def _apply_prediction_ranking(self, df: pd.DataFrame) -> pd.DataFrame:
        """Combine all available analysis metrics into a race-level prediction score."""
        if df.empty:
            return df

        df_ranked = df.copy()
        metric_configs = {
            '特徵修正勝率(%)': {'weight': 0.32, 'ascending': True},
            '勝率(%)': {'weight': 0.18, 'ascending': True},
            '綜合評分': {'weight': 0.18, 'ascending': True},
            '一致性評分': {'weight': 0.10, 'ascending': True},
            '騎師因子': {'weight': 0.05, 'ascending': True},
            '騎練合作往績因子': {'weight': 0.05, 'ascending': True},
            '較快調整後完成秒速': {'weight': 0.12, 'ascending': False},
        }

        score_series = pd.Series(0.0, index=df_ranked.index, dtype='float64')
        total_weight = 0.0

        for metric, config in metric_configs.items():
            if metric not in df_ranked.columns:
                continue

            metric_values = pd.to_numeric(df_ranked[metric], errors='coerce')
            valid_mask = metric_values.notna()
            if valid_mask.sum() == 0:
                continue

            rank_pct = metric_values.rank(
                method='average',
                pct=True,
                ascending=config['ascending']
            ) * 100.0
            score_series.loc[valid_mask] += rank_pct.loc[valid_mask] * config['weight']
            total_weight += config['weight']

        if total_weight == 0:
            df_ranked['預測得分'] = None
            df_ranked['預測排名'] = None
            return df_ranked

        score_series = score_series / total_weight
        df_ranked['預測得分'] = score_series.round(1)
        df_ranked['預測排名'] = score_series.rank(method='min', ascending=False).astype('Int64')
        return df_ranked

    def _build_predictions_sheet(self, terry_sheets: Dict[int, pd.DataFrame]) -> pd.DataFrame:
        """Build a summary sheet of the top 2 predicted horses for each race."""
        prediction_rows = []

        for race_no in sorted(terry_sheets.keys()):
            df_race = terry_sheets[race_no]
            if df_race.empty or '預測排名' not in df_race.columns:
                continue

            ranked = df_race.copy()
            ranked['預測排名'] = pd.to_numeric(ranked['預測排名'], errors='coerce')
            ranked = ranked[ranked['預測排名'].notna()].sort_values(['預測排名', '預測得分'], ascending=[True, False]).head(2)

            for _, row in ranked.iterrows():
                prediction_rows.append({
                    '場次': race_no,
                    '預測名次': int(row.get('預測排名')) if pd.notna(row.get('預測排名')) else None,
                    '馬匹編號': row.get('馬匹編號'),
                    '馬名': row.get('馬名'),
                    '騎師': row.get('騎師'),
                    '練馬師': row.get('練馬師'),
                    '預測得分': row.get('預測得分'),
                    '較快調整後完成秒速': row.get('較快調整後完成秒速'),
                    '特徵修正勝率(%)': row.get('特徵修正勝率(%)'),
                    '綜合評分': row.get('綜合評分'),
                    '一致性評分': row.get('一致性評分'),
                    '騎師因子': row.get('騎師因子'),
                    '騎練合作往績因子': row.get('騎練合作往績因子'),
                    '評估': row.get('評估'),
                })

        return pd.DataFrame(prediction_rows)

    def _build_predictions_grid_sheet(self, terry_sheets: Dict[int, pd.DataFrame], top_n: int = 4) -> pd.DataFrame:
        """Build image-style predictions layout: one race per row, horses spread horizontally."""
        grid_rows = []

        for race_no in range(1, max(terry_sheets.keys(), default=0) + 1):
            df_race = terry_sheets.get(race_no)
            row_data = {'場次': race_no}

            if df_race is not None and not df_race.empty and '預測排名' in df_race.columns:
                ranked = df_race.copy()
                ranked['預測排名'] = pd.to_numeric(ranked['預測排名'], errors='coerce')
                ranked = ranked[ranked['預測排名'].notna()].sort_values(['預測排名', '預測得分'], ascending=[True, False]).head(top_n)

                for slot, (_, horse_row) in enumerate(ranked.iterrows(), start=1):
                    row_data[f'號碼{slot}'] = horse_row.get('馬匹編號')
                    row_data[f'馬名{slot}'] = horse_row.get('馬名')

            for slot in range(1, top_n + 1):
                row_data.setdefault(f'號碼{slot}', None)
                row_data.setdefault(f'馬名{slot}', None)

            grid_rows.append(row_data)

        columns = ['場次']
        for slot in range(1, top_n + 1):
            columns.extend([f'號碼{slot}', f'馬名{slot}'])

        return pd.DataFrame(grid_rows, columns=columns)

    def _format_predictions_grid_sheet(self, worksheet, top_n: int = 4) -> None:
        """Format predictions sheet to resemble the provided image."""
        gray_fill = PatternFill(fill_type='solid', fgColor='E7E6E6')
        gold_fill = PatternFill(fill_type='solid', fgColor='FFC000')
        white_fill = PatternFill(fill_type='solid', fgColor='FFFFFF')
        thin_gray = Side(style='thin', color='BFBFBF')
        border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')
        font = Font(name='PMingLiU', size=16)

        worksheet.freeze_panes = 'A1'
        worksheet.sheet_view.showGridLines = False

        for col_idx in range(1, 2 + top_n * 2):
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            if col_idx == 1:
                worksheet.column_dimensions[col_letter].width = 8
            elif col_idx % 2 == 0:
                worksheet.column_dimensions[col_letter].width = 8
            else:
                worksheet.column_dimensions[col_letter].width = 16

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.font = font
                if cell.column == 1:
                    cell.fill = gray_fill
                    cell.alignment = center
                elif cell.column % 2 == 0:
                    cell.fill = gold_fill
                    cell.alignment = center
                else:
                    cell.fill = white_fill
                    cell.alignment = left

        for row_idx in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_idx].height = 26

    def _copy_prediction_columns_to_alfred(self, df_alfred: pd.DataFrame, df_terry: pd.DataFrame) -> pd.DataFrame:
        """Copy prediction outputs from Terry sheet into Alfred sheet for consistency."""
        if df_alfred.empty or df_terry.empty:
            return df_alfred

        merge_columns = ['馬匹編號', '預測得分', '預測排名', '綜合評分', '一致性評分', '評估', '較快調整後完成秒速']
        available_merge_columns = [col for col in merge_columns if col in df_terry.columns]
        if '馬匹編號' not in available_merge_columns:
            return df_alfred

        terry_subset = df_terry[available_merge_columns].copy()
        df_merged = df_alfred.merge(terry_subset, on='馬匹編號', how='left', suffixes=('', '_terry'))

        for col in ['預測得分', '預測排名', '綜合評分', '一致性評分', '評估', '較快調整後完成秒速']:
            terry_col = f'{col}_terry'
            if terry_col in df_merged.columns:
                df_merged[col] = df_merged[terry_col]
                df_merged = df_merged.drop(columns=[terry_col])

        return df_merged
    
    def _create_combined_files(self, date_str: str, race_count: int) -> None:
        """Create combined Excel files for different formats.
        
        Args:
            date_str: Date string in format YYYYMMDD
            race_count: Total number of races to process
        """
        race_data = {}
        terry_sheets = {}
        alfred_sheets = {}

        for race_no in range(1, race_count + 1):
            try:
                df = pd.read_excel(f"Racecard_{date_str}_{race_no}.xlsx")
                if df.empty:
                    continue
                race_data[race_no] = df

                df_terry = self._create_terry_sheet(df)
                if not df_terry.empty:
                    terry_sheets[race_no] = self._apply_prediction_ranking(df_terry)

                df_alfred = self._create_alfred_sheet(df)
                if not df_alfred.empty:
                    if race_no in terry_sheets:
                        df_alfred = self._copy_prediction_columns_to_alfred(df_alfred, terry_sheets[race_no])
                    alfred_sheets[race_no] = df_alfred
            except FileNotFoundError:
                logger.warning(f"File Racecard_{date_str}_{race_no}.xlsx not found")
            except Exception as e:
                logger.error(f"Error preparing combined data for Race {race_no}: {e}")

        predictions_df = self._build_predictions_sheet(terry_sheets)
        predictions_grid_df = self._build_predictions_grid_sheet(terry_sheets, top_n=4)

        # Standard combined file
        standard_sheets_created = 0
        with pd.ExcelWriter(f"Racecard_{date_str}.xlsx", engine='openpyxl') as writer:
            for race_no in range(1, race_count + 1):
                df = race_data.get(race_no)
                if df is not None and not df.empty:
                    df.to_excel(writer, sheet_name=f"Race_{race_no}", index=False)
                    standard_sheets_created += 1

            if not predictions_grid_df.empty:
                predictions_grid_df.to_excel(writer, sheet_name="Predictions", index=False, header=False)
                self._format_predictions_grid_sheet(writer.sheets['Predictions'], top_n=4)
            if not predictions_df.empty:
                predictions_df.to_excel(writer, sheet_name="Predictions_Detail", index=False)
        
        if standard_sheets_created == 0:
            logger.warning(f"No sheets created for standard file Racecard_{date_str}.xlsx")
        
        # Terry format
        terry_sheets_created = 0
        with pd.ExcelWriter(f"Racecard_{date_str}_Terry.xlsx", engine='openpyxl') as writer:
            for race_no in range(1, race_count + 1):
                df_terry = terry_sheets.get(race_no)
                if df_terry is not None and not df_terry.empty:
                    df_terry.to_excel(writer, sheet_name=f"Race_{race_no}", index=False)
                    terry_sheets_created += 1

            if not predictions_grid_df.empty:
                predictions_grid_df.to_excel(writer, sheet_name="Predictions", index=False, header=False)
                self._format_predictions_grid_sheet(writer.sheets['Predictions'], top_n=4)
            if not predictions_df.empty:
                predictions_df.to_excel(writer, sheet_name="Predictions_Detail", index=False)
            
            # Create a dummy sheet if no sheets were created
            if terry_sheets_created == 0:
                dummy_df = pd.DataFrame({'Message': ['No race data available']})
                dummy_df.to_excel(writer, sheet_name="No_Data", index=False)
                logger.warning("Created dummy sheet for Terry format due to no valid data")
        
        # Alfred format
        alfred_sheets_created = 0
        with pd.ExcelWriter(f"Racecard_{date_str}_Alfred.xlsx", engine='openpyxl') as writer:
            for race_no in range(1, race_count + 1):
                df_alfred = alfred_sheets.get(race_no)
                if df_alfred is not None and not df_alfred.empty:
                    df_alfred.to_excel(writer, sheet_name=f"Alfred_{race_no}", index=False)
                    alfred_sheets_created += 1

            if not predictions_grid_df.empty:
                predictions_grid_df.to_excel(writer, sheet_name="Predictions", index=False, header=False)
                self._format_predictions_grid_sheet(writer.sheets['Predictions'], top_n=4)
            if not predictions_df.empty:
                predictions_df.to_excel(writer, sheet_name="Predictions_Detail", index=False)
            
            # Create a dummy sheet if no sheets were created
            if alfred_sheets_created == 0:
                dummy_df = pd.DataFrame({'Message': ['No race data available']})
                dummy_df.to_excel(writer, sheet_name="No_Data", index=False)
                logger.warning("Created dummy sheet for Alfred format due to no valid data")
        
        logger.info(f"Created combined files: Standard({standard_sheets_created} sheets), Terry({terry_sheets_created} sheets), Alfred({alfred_sheets_created} sheets)")
    
    def download_all_racecards(self) -> None:
        """Download all racecards for the configured race date and course."""
        date_str = self.race_date.strftime('%Y%m%d')
        logger.info(f"Starting racecard download for {self.race_date} at {self.race_course}")
        
        # Dynamically detect the number of races
        race_count = self._get_race_count()
        logger.info(f"Processing {race_count} races for {self.race_date}")
        
        successful_downloads = 0
        individual_files = []  # Track individual files created
        
        # Download individual race cards (dynamically detected count)
        for race_no in range(1, race_count + 1):
            logger.info(f"Processing {self.race_date} {self.race_course} Race {race_no}")
            
            df = self._scrape_race_data(race_no)
            
            if df is not None and not df.empty:
                # Save individual race file
                filename = f"Racecard_{date_str}_{race_no}.xlsx"
                try:
                    df.to_excel(filename, index=False)
                    logger.info(f"Saved {filename}")
                    successful_downloads += 1
                    individual_files.append(filename)  # Track the file
                except Exception as e:
                    logger.error(f"Error saving {filename}: {e}")
            else:
                logger.warning(f"No data found for Race {race_no}")
        
        # Create combined files in different formats
        if successful_downloads > 0:
            try:
                self._create_combined_files(date_str, race_count)
                
                # Remove individual race files after successful combined file creation
                self._cleanup_individual_files(individual_files)
                
            except Exception as e:
                logger.error(f"Error creating combined files: {e}")
        else:
            logger.warning("No successful downloads, skipping combined file creation")
        
        logger.info(f"Racecard download completed. Successfully processed {successful_downloads} races.")
    
    def _cleanup_individual_files(self, individual_files: List[str]) -> None:
        """Remove individual race files after combined files are created."""
        removed_count = 0
        
        for filename in individual_files:
            try:
                file_path = Path(filename)
                if file_path.exists():
                    file_path.unlink()  # Delete the file
                    logger.info(f"Removed individual file: {filename}")
                    removed_count += 1
                else:
                    logger.warning(f"File not found for cleanup: {filename}")
            except Exception as e:
                logger.error(f"Error removing file {filename}: {e}")
        
        logger.info(f"Cleanup completed. Removed {removed_count} individual race files.")
    
def main():
    # Configuration
    race_date = date(2026, 4, 26)
    race_course = "ST"  # ST / HV
    
    # Create downloader and process
    downloader = HKJCRacecardDownloader(race_date, race_course)
    downloader.download_all_racecards()
    
    logger.info("Racecard download completed!")

if __name__ == "__main__":
    main()