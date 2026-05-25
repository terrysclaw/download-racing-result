import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import date
import numpy as np
import logging
from typing import Dict, List, Optional, Tuple
from pathlib import Path
import matplotlib.pyplot as plt
from matplotlib import rcParams
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class HKJCRacecardDownloader:
    def __init__(self, race_date: date, race_course: str):
        self.race_date = race_date
        self.race_course = race_course
        self.base_url = "https://racing.hkjc.com/racing/information/Chinese/Racing"
        self.df_results = self._load_historical_data()
        
    def _load_historical_data(self) -> pd.DataFrame:
        """Load historical race results."""
        try:
            df_2024 = pd.read_excel('raceresult_2024.xlsx')
            df_2025 = pd.read_excel('raceresult_2025.xlsx')
            return pd.concat([df_2024, df_2025], ignore_index=True)
        except FileNotFoundError as e:
            logger.error(f"Historical data file not found: {e}")
            return pd.DataFrame()
    
    def _initialize_data_structure(self) -> Dict[str, List]:
        """Initialize the data structure for racecard."""
        return {
            '日期': [], '馬場': [], '場次': [], '泥草': [], '賽道': [], '班次': [], '路程': [],
            '馬匹編號': [], '6次近績': [], '馬名': [], '網址': [], '烙號': [], '負磅': [],
            '騎師': [], '可能超磅': [], '檔位': [], '練馬師': [], '國際評分': [], '評分': [],
            '評分+/-': [], '排位體重': [], '排位體重+/-': [], '最佳時間': [], '馬齡': [],
            '分齡讓磅': [], '性別': [], '今季獎金': [], '優先參賽次序': [], '上賽距今日數': [],
            '配備': [], '馬主': [], '父系': [], '母系': [], '進口類別': [],
        }
    
    def _safe_int_convert(self, value: str) -> Optional[int]:
        """Safely convert string to integer."""
        try:
            return int(value.strip())
        except (ValueError, AttributeError):
            return None
    
    def _clean_text(self, text: str) -> str:
        """Clean text by removing unwanted characters."""
        return text.replace('\r', '').replace('\n', '').replace('\xa0', '').strip()
    
    def _extract_race_info(self, soup: BeautifulSoup) -> Tuple[Optional[str], Optional[str], Optional[int]]:
        """Extract race information from soup."""
        race_div = soup.find('div', class_='f_fs13')
        
        if not race_div:
            return None, None, None
            
        band = None
        race_class = None
        distance = None
        
        for col in race_div.text.split(','):
            if '賽道' in col:
                band = col.strip()
            if '班' in col:
                race_class = col.strip()
            if '米' in col:
                try:
                    distance = self._safe_int_convert(col.split('米')[0])
                except:
                    distance = None
                    
        return band, race_class, distance
    
    def _extract_horse_data(self, cells: List, race_info: Dict, data: Dict[str, List]) -> None:
        """Extract individual horse data from table cells."""
        # Add race information
        for key, value in race_info.items():
            data[key].append(value)
        
        # Extract horse-specific data with safe conversion
        data['馬匹編號'].append(self._safe_int_convert(cells[0].text))
        data['6次近績'].append(self._clean_text(cells[1].text) if len(cells) > 1 else None)
        
        # Horse name and URL
        if len(cells) > 3:
            data['馬名'].append(self._clean_text(cells[3].text))
            try:
                data['網址'].append("https://racing.hkjc.com" + cells[3].find('a')['href'])
            except:
                data['網址'].append(None)
        else:
            data['馬名'].append(None)
            data['網址'].append(None)
        
        # Continue with other fields
        fields_mapping = [
            ('烙號', 4), ('負磅', 5), ('騎師', 6), ('可能超磅', 7), ('檔位', 8),
            ('練馬師', 9), ('國際評分', 10), ('評分', 11), ('評分+/-', 12),
            ('排位體重', 13), ('排位體重+/-', 14), ('最佳時間', 15), ('馬齡', 16),
            ('分齡讓磅', 17), ('性別', 18), ('今季獎金', 19), ('優先參賽次序', 20),
            ('上賽距今日數', 21), ('配備', 22), ('馬主', 23), ('父系', 24),
            ('母系', 25), ('進口類別', 26)
        ]
        
        for field_name, index in fields_mapping:
            if index < len(cells):
                if field_name in ['負磅', '檔位', '國際評分', '評分', '排位體重', '馬齡', '分齡讓磅']:
                    data[field_name].append(self._safe_int_convert(cells[index].text))
                else:
                    data[field_name].append(self._clean_text(cells[index].text))
            else:
                data[field_name].append(None)
    
    def _get_time_adjustment_factors(self, venue: str, surface: str, distance: int) -> Dict[str, float]:
        """Get time adjustment factors based on venue, surface, and distance."""
        factors = {'weight_early': 0, 'weight_late': 0, 'gate': 0}
        
        if venue == 'HV':
            distance_factors = {
                1000: {'weight_early': 0.015, 'weight_late': 0.056, 'gate': 0.02},
                1200: {'weight_early': 0.017, 'weight_late': 0.056, 'gate': 0.033},
                1650: {'weight_early': 0.019, 'weight_late': 0.056, 'gate': 0.043},
                1800: {'weight_early': 0.021, 'weight_late': 0.056, 'gate': 0.048},
                2200: {'weight_early': 0.023, 'weight_late': 0.056, 'gate': 0.053}
            }
        else:  # ST
            if surface == '草地':
                distance_factors = {
                    1000: {'weight_early': 0.015, 'weight_late': 0.056, 'gate': -0.019},
                    1200: {'weight_early': 0.017, 'weight_late': 0.056, 'gate': 0.028},
                    1400: {'weight_early': 0.018, 'weight_late': 0.056, 'gate': 0.033},
                    1600: {'weight_early': 0.019, 'weight_late': 0.056, 'gate': 0.038},
                    1800: {'weight_early': 0.021, 'weight_late': 0.056, 'gate': 0.043},
                    2000: {'weight_early': 0.022, 'weight_late': 0.056, 'gate': 0.048},
                    2400: {'weight_early': 0.024, 'weight_late': 0.056, 'gate': 0.053}
                }
            else:  # 泥地
                distance_factors = {
                    1200: {'weight_early': 0.017, 'weight_late': 0.056, 'gate': 0.028},
                    1650: {'weight_early': 0.019, 'weight_late': 0.056, 'gate': 0.043},
                    1800: {'weight_early': 0.021, 'weight_late': 0.056, 'gate': 0.043}
                }
        
        return distance_factors.get(distance, factors)
    
    def _apply_ground_adjustment(self, time_value: float, ground_condition: str, venue: str) -> float:
        """Apply ground condition adjustment to time."""
        if ground_condition == '好地至快地':
            multiplier = 1.006 if venue == 'HV' else 1.004
            return time_value * multiplier
        elif ground_condition == '好地至黏地':
            return time_value / 1.012
        return time_value

    def _to_float(self, value) -> Optional[float]:
        """Safely convert mixed value to float."""
        if value is None or pd.isna(value) or value == '' or value == '---':
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _get_distance_scale_factor(
        self,
        tgt_venue: str, tgt_surface: str, tgt_dist: float,
        src_venue: str, src_surface: str, src_dist: float,
    ) -> float:
        """Return a time scale factor from source to target conditions.

        Uses HKJC track standard times (Class 3 benchmark) when both keys exist,
        so cross-venue and cross-distance adjustments reflect real pace differences
        rather than a naive linear distance ratio.
        Falls back to target/source distance ratio if either key is missing.
        """
        src_key = (src_venue, src_surface, int(round(src_dist)))
        tgt_key = (tgt_venue, tgt_surface, int(round(tgt_dist)))
        src_std = self._TRACK_STANDARD_TIMES.get(src_key)
        tgt_std = self._TRACK_STANDARD_TIMES.get(tgt_key)
        if src_std and tgt_std and src_std > 0:
            return tgt_std / src_std
        # Fallback: pure distance ratio (ignores venue/surface difference)
        return (tgt_dist / src_dist) if src_dist > 0 else 1.0

    def _get_section_count_by_distance(self, distance_value) -> int:
        """Map race distance to practical HK sectional count."""
        distance = self._to_float(distance_value)
        if distance is None:
            return 6
        if distance <= 1200:
            return 3
        if distance <= 1650:
            return 4
        if distance <= 2000:
            return 5
        return 6

    def _get_section_sizes(self, distance: float) -> List[int]:
        """Return HKJC section sizes (metres, start→finish) for the given race distance.

        HKJC measures sections from fixed finish-relative timing marks at every 400 m
        (400 m, 800 m, 1200 m, 1600 m, 2000 m from finish).  The *last* section is
        always 400 m; intermediate sections are also 400 m; the *first* section is
        whatever remains (e.g. 200 m for 1400 m, 450 m for 1650 m).
        """
        d = int(round(distance))
        n = self._get_section_count_by_distance(d)
        trailing = (n - 1) * 400          # metres covered by all sections except the first
        first = d - trailing
        return [first] + [400] * (n - 1)

    # Surface time conversion factors: fallback when standard-time lookup fails.
    # Based on HKJC table: ST 1200m turf=69.00s vs AW=68.95s → near-parity.
    _SURFACE_FACTORS: Dict[Tuple[str, str], float] = {
        ('草地', '泥地'): 0.999,
        ('泥地', '草地'): 1.001,
    }

    # HKJC official track standard times (seconds), Class 3 benchmark, updated 2025-08-26.
    # Used to compute accurate cross-distance / cross-venue scaling factors.
    _TRACK_STANDARD_TIMES: Dict[Tuple[str, str, int], float] = {
        # Sha Tin Turf (沙田草地)
        ('ST', '草地', 1000): 56.45,
        ('ST', '草地', 1200): 69.00,
        ('ST', '草地', 1400): 81.65,
        ('ST', '草地', 1600): 94.70,
        ('ST', '草地', 1800): 107.50,
        ('ST', '草地', 2000): 121.90,
        ('ST', '草地', 2400): 147.00,
        # Happy Valley Turf (跑馬地草地)
        ('HV', '草地', 1000): 56.65,
        ('HV', '草地', 1200): 69.60,
        ('HV', '草地', 1650): 99.90,
        ('HV', '草地', 1800): 109.45,
        ('HV', '草地', 2200): 136.60,
        # Sha Tin All-Weather / Dirt (沙田全天候)
        ('ST', '泥地', 1200): 68.95,
        ('ST', '泥地', 1650): 99.05,
        ('ST', '泥地', 1800): 108.05,
    }

    # HKJC Class-3 reference sectional times (seconds, start→finish) from
    # https://racing.hkjc.com/zh-hk/local/page/racing-course-time (updated 2025-08-26).
    # Used for reference-based cross-distance section remapping so that section
    # sizes and pace profiles are matched correctly (e.g. 1400m sec1=200m vs
    # 1600m sec1=400m are converted via their respective reference values).
    _TRACK_REFERENCE_SECTIONALS: Dict[Tuple[str, str, int], Tuple[float, ...]] = {
        # Sha Tin Turf
        ('ST', '草地', 1000): (13.05, 20.65, 22.75),
        ('ST', '草地', 1200): (23.70, 22.35, 22.95),
        ('ST', '草地', 1400): (13.45, 21.80, 23.15, 23.25),
        ('ST', '草地', 1600): (24.50, 22.90, 23.80, 23.50),
        ('ST', '草地', 1800): (13.85, 22.30, 23.80, 24.00, 23.55),
        ('ST', '草地', 2000): (26.05, 24.70, 24.05, 23.55, 23.55),
        ('ST', '草地', 2400): (25.60, 24.50, 25.35, 23.85, 23.75, 23.95),  # 分級賽
        # Happy Valley Turf
        ('HV', '草地', 1000): (12.50, 21.00, 23.15),
        ('HV', '草地', 1200): (23.50, 22.55, 23.55),
        ('HV', '草地', 1650): (27.95, 23.85, 24.25, 23.85),
        ('HV', '草地', 1800): (13.75, 23.00, 24.30, 24.45, 23.95),
        ('HV', '草地', 2200): (14.35, 23.70, 24.95, 24.85, 24.45, 24.30),
        # Sha Tin All-Weather
        ('ST', '泥地', 1200): (23.30, 22.05, 23.20),
        ('ST', '泥地', 1650): (27.90, 23.00, 23.75, 23.95),
        ('ST', '泥地', 1800): (13.75, 22.80, 23.70, 23.85, 23.95),
    }

    def _get_recent_three_runs(self, row: pd.Series) -> Tuple[pd.DataFrame, Optional[float], int]:
        """Get recent runs with progressive fallback for horses without same-course history.

        Returns:
            (DataFrame, source_distance, fallback_level)
            fallback_level: 1=same venue/surface/dist, 2=same surface/dist any venue,
                            3=same surface nearby dist (±250m), 4=same surface any dist,
                            5=any surface, 0=no data
        """
        if self.df_results.empty:
            return pd.DataFrame(), None, 0

        brand = row['烙號']
        surface = row['泥草']
        target_dist_raw = row['路程']
        target_dist = self._to_float(target_dist_raw)

        # Level 1: Exact match (same venue, surface, distance) — original behaviour
        exact_filter = (
            (self.df_results['布號'] == brand) &
            (self.df_results['馬場'] == row['馬場']) &
            (self.df_results['泥草'] == surface) &
            (self.df_results['路程'] == target_dist_raw) &
            (self.df_results['獨贏賠率'] > 0)
        )
        result = self.df_results[exact_filter].tail(3)
        if not result.empty:
            return result, target_dist, 1

        # Level 2: Same surface + distance, any venue (e.g. ST record for HV race)
        same_dist_filter = (
            (self.df_results['布號'] == brand) &
            (self.df_results['泥草'] == surface) &
            (self.df_results['路程'] == target_dist_raw) &
            (self.df_results['獨贏賠率'] > 0)
        )
        result = self.df_results[same_dist_filter].tail(3)
        if not result.empty:
            return result, target_dist, 2

        # Level 3: Same surface, nearby distance (±200m), any venue
        if target_dist is not None:
            numeric_dist = pd.to_numeric(self.df_results['路程'], errors='coerce')
            nearby_filter = (
                (self.df_results['布號'] == brand) &
                (self.df_results['泥草'] == surface) &
                ((numeric_dist - target_dist).abs() <= 250) &
                (self.df_results['獨贏賠率'] > 0)
            )
            result = self.df_results[nearby_filter].tail(3)
            if not result.empty:
                source_dist = pd.to_numeric(result['路程'], errors='coerce').mean()
                return result, float(source_dist) if pd.notna(source_dist) else target_dist, 3

        # Level 4: Same surface, any distance
        any_dist_filter = (
            (self.df_results['布號'] == brand) &
            (self.df_results['泥草'] == surface) &
            (self.df_results['獨贏賠率'] > 0)
        )
        result = self.df_results[any_dist_filter].tail(3)
        if not result.empty:
            source_dist = pd.to_numeric(result['路程'], errors='coerce').mean()
            return result, float(source_dist) if pd.notna(source_dist) else target_dist, 4

        # Level 5: Any surface (last resort)
        any_filter = (
            (self.df_results['布號'] == brand) &
            (self.df_results['獨贏賠率'] > 0)
        )
        result = self.df_results[any_filter].tail(3)
        if not result.empty:
            source_dist = pd.to_numeric(result['路程'], errors='coerce').mean()
            return result, float(source_dist) if pd.notna(source_dist) else target_dist, 5

        return pd.DataFrame(), None, 0

    def _calculate_predicted_sections(
        self,
        row: pd.Series,
        avg_sections: Dict[int, Optional[float]],
        recent_runs: pd.DataFrame,
        target_section_count: int,
        jt_rate: Optional[float] = None
    ) -> Dict[int, Optional[float]]:
        """Estimate today's sectional times using weight, draw, and jockey-trainer adjustments."""
        predicted = {i: (avg_sections.get(i) if i <= target_section_count else None) for i in range(1, 7)}

        valid_section_count = len([avg_sections.get(i) for i in range(1, target_section_count + 1) if avg_sections.get(i) is not None])
        if valid_section_count == 0:
            return predicted

        distance = self._to_float(row.get('路程'))
        if not distance or distance <= 0:
            return predicted

        current_weight = self._to_float(row.get('負磅'))
        current_gate = self._to_float(row.get('檔位'))
        hist_weight_mean = pd.to_numeric(recent_runs.get('實際負磅'), errors='coerce').mean()
        hist_gate_mean = pd.to_numeric(recent_runs.get('檔位'), errors='coerce').mean()

        weight_diff = 0.0
        gate_diff = 0.0
        if current_weight is not None and pd.notna(hist_weight_mean):
            weight_diff = current_weight - float(hist_weight_mean)
        if current_gate is not None and pd.notna(hist_gate_mean):
            gate_diff = current_gate - float(hist_gate_mean)

        factors = self._get_time_adjustment_factors(
            row['馬場'], row['泥草'], int(distance)
        )

        section_distance = distance / valid_section_count
        early_adj_per_section = weight_diff * factors['weight_early'] / distance * section_distance
        late_adj_per_section = weight_diff * factors['weight_late'] / distance * section_distance

        for i in range(1, target_section_count + 1):
            base_time = avg_sections.get(i)
            if base_time is None:
                continue

            # Gate impact is concentrated in the first 2 sections.
            if i == 1:
                gate_adj = gate_diff * factors['gate'] * 0.7
            elif i == 2:
                gate_adj = gate_diff * factors['gate'] * 0.3
            else:
                gate_adj = 0.0

            half_point = max(1, valid_section_count // 2)
            weight_adj = early_adj_per_section if i <= half_point else late_adj_per_section
            predicted[i] = round(base_time + weight_adj + gate_adj, 3)

        # Jockey-trainer adjustment: pairs above/below 33% baseline run proportionally faster/slower.
        # Coefficient 0.02 means a 43% pair (~+10% above baseline) adjusts each section by ~-0.20%.
        if jt_rate is not None:
            jt_factor = 1.0 - (min(jt_rate, 0.70) - 0.33) * 0.02
            for i in range(1, target_section_count + 1):
                if predicted[i] is not None:
                    predicted[i] = round(predicted[i] * jt_factor, 3)

        return predicted

    def _add_win_probability(self, df: pd.DataFrame) -> pd.DataFrame:
        """Estimate win probability by race using predicted total time."""
        if df.empty or '預估總段速' not in df.columns:
            return df

        if '勝出機率' not in df.columns:
            df['勝出機率'] = None
        if '公平賠率(扣稅)' not in df.columns:
            df['公平賠率(扣稅)'] = None

        # HKJC takeout rate for Win pool; fair odds are reduced after deduction.
        hkjc_tax_rate = 0.175

        for race_no, race_df in df.groupby('場次'):
            race_index = race_df.index
            total_times = pd.to_numeric(df.loc[race_index, '預估總段速'], errors='coerce')
            valid_times = total_times.dropna()

            if len(race_index) == 0:
                continue

            # If no horse has a predicted total, assign equal chance to avoid hard exclusion.
            if valid_times.empty:
                uniform_prob = round(100.0 / len(race_index), 2)
                for idx in race_index:
                    df.at[idx, '勝出機率'] = uniform_prob
                    prob_decimal = uniform_prob / 100.0
                    fair_odds = ((1.0 - hkjc_tax_rate) / prob_decimal) if prob_decimal > 0 else np.nan
                    df.at[idx, '公平賠率(扣稅)'] = round(float(fair_odds), 2) if pd.notna(fair_odds) else None
                continue

            # Convert faster (smaller) times into probabilities using a softmax-like transform.
            # Temperature adapts to field spread; lower temperature increases separation.
            spread = valid_times.std()
            temperature = max(float(spread), 0.15) if pd.notna(spread) else 0.25
            best_time = valid_times.min()

            # Missing totals: use rating-based interpolation when possible, else conservative penalty.
            if total_times.isna().any():
                worst_time = valid_times.max()
                missing_penalty = max(0.4, (float(spread) * 0.8) if pd.notna(spread) else 0.4)
                default_imputed = float(worst_time) + missing_penalty
                all_ratings = pd.to_numeric(df.loc[race_index, '評分'], errors='coerce')
                rated_valid_idx = valid_times.index.intersection(all_ratings.dropna().index)
                model_times = total_times.copy().astype(float)
                for idx in total_times[total_times.isna()].index:
                    horse_rating = all_ratings.get(idx)
                    if pd.notna(horse_rating) and len(rated_valid_idx) >= 2:
                        rated_times = valid_times.reindex(rated_valid_idx).dropna()
                        rated_ratings = all_ratings.reindex(rated_valid_idx).dropna()
                        rating_min, rating_max = float(rated_ratings.min()), float(rated_ratings.max())
                        rating_range = rating_max - rating_min
                        if rating_range > 0:
                            pct = (float(horse_rating) - rating_min) / rating_range
                            # Higher rating (larger pct) → faster (smaller) time
                            estimated = float(rated_times.max()) - pct * (float(rated_times.max()) - float(rated_times.min()))
                            model_times.at[idx] = max(float(best_time) * 0.98, min(estimated, default_imputed))
                        else:
                            model_times.at[idx] = default_imputed
                    else:
                        model_times.at[idx] = default_imputed
            else:
                model_times = total_times

            score = np.exp(-(model_times - best_time) / temperature)
            score_sum = score.sum()
            if score_sum <= 0:
                continue

            prob = score / score_sum
            for idx, p in prob.items():
                prob_pct = round(float(p) * 100, 2)
                df.at[idx, '勝出機率'] = prob_pct
                fair_odds = ((1.0 - hkjc_tax_rate) / float(p)) if float(p) > 0 else np.nan
                df.at[idx, '公平賠率(扣稅)'] = round(float(fair_odds), 2) if pd.notna(fair_odds) else None

        return df

    def _simulate_race_shape(self, df: pd.DataFrame) -> pd.DataFrame:
        """Simulate likely race positioning from forecast early pace and draw."""
        if df.empty or '預估頭段(第1+2段)' not in df.columns:
            return df

        df['預計各分段排名'] = None
        df['模擬走勢'] = None
        df['模擬走位'] = None

        for race_no, race_df in df.groupby('場次'):
            valid_index = race_df[race_df['預估頭段(第1+2段)'].notna()].index
            if len(valid_index) == 0:
                continue

            # Build per-section cumulative ranking sequence, e.g. 2-2-1-1.
            rank_sequences = {idx: [] for idx in valid_index}
            cumulative_times = pd.Series(0.0, index=valid_index, dtype='float64')

            for section_no in range(1, 7):
                section_col = f'預估今場第 {section_no} 段'
                section_values = pd.to_numeric(df.loc[valid_index, section_col], errors='coerce')

                valid_mask = section_values.notna() & cumulative_times.notna()
                if not valid_mask.any():
                    continue

                cumulative_times.loc[valid_mask] = cumulative_times.loc[valid_mask] + section_values.loc[valid_mask]
                cumulative_times.loc[~valid_mask] = np.nan

                section_rank = cumulative_times.rank(method='min', ascending=True)
                for idx in valid_index:
                    rank_value = section_rank.loc[idx]
                    if pd.notna(rank_value):
                        rank_sequences[idx].append(str(int(rank_value)))

            for idx in valid_index:
                if rank_sequences[idx]:
                    df.at[idx, '預計各分段排名'] = '-'.join(rank_sequences[idx])

            early_rank = df.loc[valid_index, '預估頭段(第1+2段)'].rank(method='min', ascending=True)
            field_size = len(valid_index)

            for idx in valid_index:
                rank_value = int(early_rank.loc[idx])
                gate = self._to_float(df.at[idx, '檔位'])

                if rank_value <= max(1, int(np.ceil(field_size * 0.3))):
                    pace_bucket = '前置'
                elif rank_value <= max(2, int(np.ceil(field_size * 0.7))):
                    pace_bucket = '中置'
                else:
                    pace_bucket = '後上'

                if gate is None:
                    gate_bucket = '不明檔位'
                elif gate <= 4:
                    gate_bucket = '內檔'
                elif gate <= 8:
                    gate_bucket = '中檔'
                else:
                    gate_bucket = '外檔'

                if pace_bucket == '前置' and gate_bucket == '內檔':
                    style = '前置貼欄'
                elif pace_bucket == '前置' and gate_bucket == '外檔':
                    style = '外疊搶放'
                elif pace_bucket == '中置' and gate_bucket == '內檔':
                    style = '包廂跟前'
                elif pace_bucket == '後上' and gate_bucket == '外檔':
                    style = '留後外疊'
                elif pace_bucket == '後上':
                    style = '留後追趕'
                else:
                    style = f'{pace_bucket}守位'

                df.at[idx, '模擬走勢'] = style
                df.at[idx, '模擬走位'] = f"{style} ({gate_bucket}, 頭段排名{rank_value})"

        return df

    def _get_jockey_trainer_pair(self, jockey: str, trainer: str) -> Optional[pd.DataFrame]:
        """Return historical rows for a jockey-trainer pair in the last 3 months, stripping overweight notation."""
        import re
        from datetime import timedelta
        jockey = re.sub(r'\s*\([^)]*\)', '', str(jockey or '')).strip()
        trainer = re.sub(r'\s*\([^)]*\)', '', str(trainer or '')).strip()
        if self.df_results.empty or not jockey or not trainer:
            return None
        cutoff = pd.Timestamp(self.race_date) - pd.DateOffset(months=3)
        date_col = '日期' if '日期' in self.df_results.columns else None
        pair_mask = (
            (self.df_results['騎師'] == jockey) &
            (self.df_results['練馬師'] == trainer)
        )
        if date_col:
            dates = pd.to_datetime(self.df_results[date_col], errors='coerce')
            pair_mask = pair_mask & (dates >= cutoff)
        pair = self.df_results[pair_mask]
        return pair if not pair.empty else None

    def _calculate_jockey_trainer_rate(self, jockey: str, trainer: str) -> Optional[str]:
        """Return jockey-trainer place rate (top 3) as formatted string."""
        pair = self._get_jockey_trainer_pair(jockey, trainer)
        if pair is None:
            return None
        total = len(pair)
        placed = pair['名次'].apply(lambda x: pd.to_numeric(x, errors='coerce')).le(3).sum()
        rate = placed / total
        return f"{rate:.0%}({int(placed)}/{total})"

    def _calculate_jockey_trainer_rate_numeric(self, jockey: str, trainer: str) -> Optional[float]:
        """Return jockey-trainer place rate (top 3) as float (0.0–1.0), or None."""
        pair = self._get_jockey_trainer_pair(jockey, trainer)
        if pair is None:
            return None
        total = len(pair)
        placed = pair['名次'].apply(lambda x: pd.to_numeric(x, errors='coerce')).le(3).sum()
        return placed / total

    def _remap_sections_by_finish_distance(
        self,
        src_sections: Dict[int, Optional[float]],
        src_dist: float,
        tgt_dist: float,
        time_scale: float,
    ) -> Dict[int, Optional[float]]:
        """Remap source sectional times to the target section layout.

        HKJC times sections from fixed finish-relative marks (400 m, 800 m, …).
        Sections that cover the *same finish-relative range* are matched directly;
        any extra metres at the *start* (when tgt_dist > src_dist) are estimated
        using the pace of the first known source section.

        The whole result is normalised so that Σ target_sections = src_total × time_scale,
        removing any error introduced by the start-section extrapolation.
        """
        src_sizes = self._get_section_sizes(src_dist)
        tgt_sizes = self._get_section_sizes(tgt_dist)
        src_n = len(src_sizes)
        tgt_n = len(tgt_sizes)

        # Pace (s/m) for each known source section; fill gaps with mean.
        src_paces: Dict[int, float] = {}
        for i in range(1, src_n + 1):
            v = src_sections.get(i)
            if v is not None and src_sizes[i - 1] > 0:
                src_paces[i] = float(v) / src_sizes[i - 1]
        if not src_paces:
            return {}
        avg_pace = float(np.mean(list(src_paces.values())))
        for i in range(1, src_n + 1):
            src_paces.setdefault(i, avg_pace)
        first_src_pace = src_paces[1]   # used to extrapolate "before source start"

        # Cumulative distance from finish to the *start* of each section (1-indexed).
        # For source: src_start_ff[0] = src_dist (start of sec 1 = src_dist from finish)
        #             src_start_ff[i] = src_dist - sum(src_sizes[0..i])
        def _start_from_finish(sizes: List[int]) -> List[float]:
            result = []
            cum = float(sum(sizes))
            for s in sizes:
                result.append(cum)
                cum -= s
            return result   # result[i] = dist from finish to start of section i+1

        src_sff = _start_from_finish(src_sizes)   # index 0 = sec1, etc.
        tgt_sff = _start_from_finish(tgt_sizes)

        raw: Dict[int, float] = {}
        for j in range(tgt_n):
            t_far  = tgt_sff[j]                       # farther from finish
            t_near = t_far - tgt_sizes[j]              # closer to finish

            t_total = 0.0
            remaining = float(tgt_sizes[j])

            for i in range(src_n):
                s_far  = src_sff[i]
                s_near = s_far - src_sizes[i]

                ov_far  = min(t_far,  s_far)
                ov_near = max(t_near, s_near)
                if ov_far <= ov_near:
                    continue

                t_total   += src_paces[i + 1] * (ov_far - ov_near)
                remaining -= (ov_far - ov_near)

            # Any remaining metres are *before* the source start → extrapolate.
            if remaining > 1e-9:
                t_total += first_src_pace * remaining

            raw[j + 1] = t_total

        # Normalise: ensure Σ raw × norm = src_total × time_scale.
        known_src = [float(v) for v in src_sections.values() if v is not None]
        src_total = sum(known_src) if known_src else 0.0
        raw_total = sum(raw.values())
        if raw_total > 1e-9 and src_total > 0:
            norm = (src_total * time_scale) / raw_total
        else:
            norm = time_scale

        return {j: round(t * norm, 3) for j, t in raw.items()}

    def _remap_sections_reference_based(
        self,
        src_sections: Dict[int, Optional[float]],
        src_dist: float,
        src_venue: str,
        src_surface: str,
        tgt_dist: float,
        tgt_venue: str,
        tgt_surface: str,
        time_scale: float,
    ) -> Dict[int, Optional[float]]:
        """Remap source sectional times to target using HKJC Class-3 reference sectionals.

        Algorithm (all sections are tail-aligned, i.e. k=1 = last 400m):
          1. Compute ratio_k = horse_src[k] / ref_src[k] for each common tail section.
          2. pred_tgt[k] = ref_tgt[k] × ratio_k  (direct calibration).
          3. For extra target sections at the START (tgt_dist > src_dist, no source
             counterpart), use avg_ratio so the horse's overall performance level is
             preserved.
          4. Normalise so Σ result = src_total × time_scale.

        Falls back to pace-based remapping when reference data is unavailable.
        """
        src_key = (src_venue, src_surface, int(round(src_dist)))
        tgt_key = (tgt_venue, tgt_surface, int(round(tgt_dist)))
        ref_src = self._TRACK_REFERENCE_SECTIONALS.get(src_key)
        ref_tgt = self._TRACK_REFERENCE_SECTIONALS.get(tgt_key)

        if not ref_src or not ref_tgt:
            return self._remap_sections_by_finish_distance(
                src_sections, src_dist, tgt_dist, time_scale
            )

        src_n = len(ref_src)
        tgt_n = len(ref_tgt)

        # Compute horse's performance ratio vs reference, tail-aligned.
        # k=1 → last section, k=2 → second-to-last, etc.
        ratios: Dict[int, float] = {}
        for k in range(1, src_n + 1):
            src_idx = src_n + 1 - k          # 1-based section index
            v = src_sections.get(src_idx)
            ref_v = float(ref_src[src_idx - 1])
            if v is not None and ref_v > 0:
                ratios[k] = float(v) / ref_v

        if not ratios:
            return self._remap_sections_by_finish_distance(
                src_sections, src_dist, tgt_dist, time_scale
            )

        avg_ratio = float(np.mean(list(ratios.values())))

        # Build target sections: pred_tgt[j] = ref_tgt[j] × ratio_k (tail-aligned).
        raw: Dict[int, float] = {}
        for j in range(1, tgt_n + 1):
            k = tgt_n + 1 - j               # tail-alignment index for this target section
            ratio = ratios.get(k, avg_ratio) # avg_ratio for extra start sections
            raw[j] = float(ref_tgt[j - 1]) * ratio

        # Normalise so total = src_total × time_scale.
        known_src = [float(v) for v in src_sections.values() if v is not None]
        src_total = sum(known_src) if known_src else 0.0
        raw_total = sum(raw.values())
        norm = (src_total * time_scale) / raw_total if raw_total > 1e-9 and src_total > 0 else 1.0

        return {j: round(t * norm, 3) for j, t in raw.items()}

    def _compute_head_tail_800(
        self,
        predicted_sections: Dict[int, Optional[float]],
        total_dist: Optional[float],
        section_count: int,
    ) -> Tuple[Optional[float], Optional[float]]:
        """Compute first-800m and last-800m predicted times using HKJC section sizes.

        Uses actual HKJC section sizes (e.g. [200,400,400,400] for 1400m) so that
        head/tail exactly cover 800 m of course rather than an arbitrary fraction.
        """
        if not total_dist or total_dist <= 0 or section_count <= 0:
            return None, None

        section_sizes = self._get_section_sizes(total_dist)
        if len(section_sizes) != section_count:
            # Fallback: equal sections
            sz = total_dist / section_count
            section_sizes = [sz] * section_count

        target = min(800.0, total_dist)

        def _accumulate(from_start: bool) -> Optional[float]:
            remaining = target
            acc = 0.0
            pairs = (
                list(zip(range(1, section_count + 1), section_sizes))
                if from_start
                else list(zip(range(section_count, 0, -1), reversed(section_sizes)))
            )
            for idx, sz in pairs:
                t = predicted_sections.get(idx)
                if t is None:
                    return None
                if remaining >= sz - 1e-9:
                    acc += float(t)
                    remaining -= sz
                else:
                    acc += float(t) * (remaining / sz)
                    remaining = 0.0
                if remaining <= 1e-9:
                    break
            return round(acc, 3) if remaining <= 1e-9 else None

        return _accumulate(True), _accumulate(False)

    def _add_sectional_forecast(self, df: pd.DataFrame) -> pd.DataFrame:
        """Add 3-run sectional averages and forecast sectionals for today's race."""
        avg_cols = [f'近3仗第 {i} 段平均' for i in range(1, 7)]
        pred_cols = [f'預估今場第 {i} 段' for i in range(1, 7)]
        extra_cols = ['近3仗樣本數', '近況來源', '預估頭段(第1+2段)', '預估末段(最後2段)', '預估末段(第5+6段)', '預估總段速', '勝出機率', '公平賠率(扣稅)', '預計各分段排名', '模擬走勢', '模擬走位', '騎練合作上名率']

        for col in avg_cols + pred_cols + extra_cols:
            df[col] = None

        _FALLBACK_LABELS = {1: '同場同程', 2: '換場同程', 3: '換距近', 4: '換距遠', 5: '換面'}

        for index, row in df.iterrows():
            recent_runs, source_dist, fallback_level = self._get_recent_three_runs(row)
            if recent_runs.empty:
                continue

            target_section_count = self._get_section_count_by_distance(row.get('路程'))
            sample_count = len(recent_runs)
            avg_sections = {}
            has_ground_data = '場地狀況' in recent_runs.columns and '馬場' in recent_runs.columns
            target_dist = self._to_float(row.get('路程'))
            tgt_venue_str = str(row.get('馬場', ''))
            tgt_surface_str = str(row.get('泥草', ''))

            if fallback_level >= 3 and target_dist is not None:
                # Levels 3-5: runs may come from MIXED distances (e.g. 1600m + 1800m + 1800m).
                # Section 1 of a 1800m race is only 200 m while section 1 of a 1600m race is
                # 400 m — naively averaging them produces nonsense.  Instead, remap each run
                # individually to the target distance first, then average the remapped values.
                remapped_runs_list: List[Dict[int, Optional[float]]] = []
                for _, run_row in recent_runs.iterrows():
                    run_dist = self._to_float(run_row.get('路程'))
                    if run_dist is None:
                        continue
                    run_sections_raw: Dict[int, Optional[float]] = {}
                    for i in range(1, 7):
                        v = self._to_float(run_row.get(f'第 {i} 段'))
                        if v is not None and has_ground_data:
                            ground = str(run_row.get('場地狀況', ''))
                            venue = str(run_row.get('馬場', ''))
                            v = self._apply_ground_adjustment(v, ground, venue)
                        run_sections_raw[i] = v
                    run_venue = str(run_row.get('馬場', tgt_venue_str))
                    run_surface = str(run_row.get('泥草', tgt_surface_str))
                    scale = self._get_distance_scale_factor(
                        tgt_venue_str, tgt_surface_str, target_dist,
                        run_venue, run_surface, run_dist,
                    )
                    remapped = self._remap_sections_reference_based(
                        run_sections_raw, run_dist, run_venue, run_surface,
                        target_dist, tgt_venue_str, tgt_surface_str, scale,
                    )
                    if remapped:
                        remapped_runs_list.append(remapped)
                for i in range(1, 7):
                    if remapped_runs_list:
                        vals = [r.get(i) for r in remapped_runs_list if r.get(i) is not None]
                        avg_sections[i] = round(float(np.mean(vals)), 3) if vals else None
                    else:
                        avg_sections[i] = None
                    df.at[index, f'近3仗第 {i} 段平均'] = avg_sections[i] if i <= target_section_count else None
            else:
                # Levels 1-2: all runs are at the same distance — average sections directly.
                # For level 2 (same distance, different venue), apply a venue/surface time
                # scale to the averaged sections afterwards.
                for i in range(1, 7):
                    section_series = pd.to_numeric(recent_runs.get(f'第 {i} 段'), errors='coerce')
                    if has_ground_data:
                        adjusted_values = []
                        for idx, val in section_series.items():
                            if pd.notna(val):
                                ground = str(recent_runs.at[idx, '場地狀況'])
                                venue = str(recent_runs.at[idx, '馬場'])
                                adjusted_values.append(self._apply_ground_adjustment(val, ground, venue))
                        avg_value = np.mean(adjusted_values) if adjusted_values else np.nan
                    else:
                        avg_value = section_series.mean()
                    avg_sections[i] = round(float(avg_value), 3) if pd.notna(avg_value) else None
                    df.at[index, f'近3仗第 {i} 段平均'] = avg_sections[i] if i <= target_section_count else None

                if fallback_level == 2 and target_dist is not None:
                    src_venue = str(recent_runs['馬場'].mode().iloc[0]) if '馬場' in recent_runs.columns else tgt_venue_str
                    src_surface = str(recent_runs['泥草'].mode().iloc[0]) if '泥草' in recent_runs.columns else tgt_surface_str
                    eff_src_dist = source_dist if source_dist is not None else target_dist
                    if eff_src_dist is not None:
                        scale = self._get_distance_scale_factor(
                            tgt_venue_str, tgt_surface_str, target_dist,
                            src_venue, src_surface, eff_src_dist,
                        )
                        remapped = self._remap_sections_reference_based(
                            avg_sections, eff_src_dist, src_venue, src_surface,
                            target_dist, tgt_venue_str, tgt_surface_str, scale,
                        )
                        if remapped:
                            avg_sections = remapped
                            for i in range(1, target_section_count + 1):
                                df.at[index, f'近3仗第 {i} 段平均'] = avg_sections.get(i)

            jt_rate = self._calculate_jockey_trainer_rate_numeric(
                str(row.get('騎師', '') or '').strip(),
                str(row.get('練馬師', '') or '').strip()
            )
            predicted_sections = self._calculate_predicted_sections(
                row,
                avg_sections,
                recent_runs,
                target_section_count,
                jt_rate=jt_rate
            )
            for i in range(1, 7):
                df.at[index, f'預估今場第 {i} 段'] = predicted_sections.get(i)

            valid_all = [predicted_sections.get(i) for i in range(1, target_section_count + 1)
                         if predicted_sections.get(i) is not None]
            race_dist = self._to_float(row.get('路程'))
            head_800, tail_800 = self._compute_head_tail_800(
                predicted_sections, race_dist, target_section_count
            )

            df.at[index, '近3仗樣本數'] = sample_count
            df.at[index, '近況來源'] = _FALLBACK_LABELS.get(fallback_level, '-')
            df.at[index, '預估頭段(第1+2段)'] = head_800
            df.at[index, '預估末段(最後2段)'] = tail_800
            df.at[index, '預估末段(第5+6段)'] = tail_800
            df.at[index, '預估總段速'] = round(sum(valid_all), 3) if valid_all else None

        df['騎練合作上名率'] = df.apply(
            lambda row: self._calculate_jockey_trainer_rate(
                str(row.get('騎師', '') or '').strip(),
                str(row.get('練馬師', '') or '').strip()
            ), axis=1
        )

        df = self._add_win_probability(df)
        return self._simulate_race_shape(df)
    
    def _scrape_race_data(self, race_no: int) -> Optional[pd.DataFrame]:
        """Scrape race data for a specific race."""
        url = f"{self.base_url}/RaceCard.aspx?RaceDate={self.race_date.strftime('%Y/%m/%d')}&Racecourse={self.race_course}&RaceNo={race_no}"
        
        try:
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Extract race information
            band, race_class, distance = self._extract_race_info(soup)
            
            race_info = {
                '日期': self.race_date,
                '馬場': self.race_course,
                '場次': race_no,
                '泥草': '草地' if band else '泥地',
                '賽道': band,
                '班次': race_class,
                '路程': distance
            }
            
            # Find racecard table
            racecard_table = soup.find('table', id='racecardlist')
            if not racecard_table:
                logger.warning(f"No racecard table found for Race {race_no}")
                return None
            
            # Extract horse data
            data = self._initialize_data_structure()
            tbody = racecard_table.find('tbody')
            
            if tbody:
                for tr in tbody.find_all('tr')[2:]:  # Skip header rows
                    cells = tr.find_all('td')
                    if len(cells) >= 4:  # Ensure minimum required cells
                        self._extract_horse_data(cells, race_info, data)
            
            df = pd.DataFrame(data)
            
            if not df.empty:
                df = self._add_sectional_forecast(df)
            
            return df
            
        except requests.RequestException as e:
            logger.error(f"Error fetching race {race_no}: {e}")
            return None
        except Exception as e:
            logger.error(f"Unexpected error processing race {race_no}: {e}")
            return None
    
    def _create_terry_sheet(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create Terry AI formatted sheet."""
        required_columns = [
            '場次', '班次', '路程', '馬匹編號', '馬名', '騎師', '練馬師', '騎練合作上名率', '評分',
            '負磅', '檔位', '最佳時間',
            '近3仗樣本數', '近況來源',
            '近3仗第 1 段平均', '近3仗第 2 段平均', '近3仗第 3 段平均',
            '近3仗第 4 段平均', '近3仗第 5 段平均', '近3仗第 6 段平均',
            '預估今場第 1 段', '預估今場第 2 段', '預估今場第 3 段',
            '預估今場第 4 段', '預估今場第 5 段', '預估今場第 6 段',
            '預估頭段(第1+2段)', '預估末段(最後2段)', '預估末段(第5+6段)', '預估總段速', '勝出機率', '公平賠率(扣稅)', '預計各分段排名', '模擬走勢', '模擬走位',
            '網址'
        ]

        existing_columns = [col for col in required_columns if col in df.columns]
        df_terry = df[existing_columns].copy()
        return df_terry

    def _apply_fallback_italic(self, writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
        """Apply italic font to prediction columns for horses whose data come from
        a cross-distance or cross-venue fallback (近況來源 != '同場同程')."""
        if '近況來源' not in df.columns:
            return
        italic_cols = (
            [f'近3仗第 {i} 段平均' for i in range(1, 7)] +
            [f'預估今場第 {i} 段' for i in range(1, 7)] +
            ['預估頭段(第1+2段)', '預估末段(最後2段)', '預估末段(第5+6段)', '預估總段速']
        )
        existing_italic_cols = [c for c in italic_cols if c in df.columns]
        if not existing_italic_cols:
            return
        ws = writer.book[sheet_name]
        for row_idx, source_val in enumerate(df['近況來源'].tolist(), start=2):
            if str(source_val) == '同場同程':
                continue
            for col_name in existing_italic_cols:
                col_idx = df.columns.get_loc(col_name) + 1
                cell = ws.cell(row=row_idx, column=col_idx)
                existing_font = cell.font
                cell.font = Font(
                    name=existing_font.name,
                    size=existing_font.size,
                    bold=existing_font.bold,
                    italic=True,
                    color=existing_font.color,
                )

    def _highlight_fastest_section_cells(self, writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
        """Highlight the fastest value in each predicted section column with yellow fill."""
        section_cols = [f'預估今場第 {i} 段' for i in range(1, 7)] + ['預估頭段(第1+2段)', '預估末段(最後2段)']
        existing_cols = [col for col in section_cols if col in df.columns]
        if not existing_cols:
            return

        ws = writer.book[sheet_name]
        orange_fill = PatternFill(fill_type='solid', fgColor='FF6600')
        yellow_fill = PatternFill(fill_type='solid', fgColor='FFFF00')

        for col_name in existing_cols:
            section_values = pd.to_numeric(df[col_name], errors='coerce')
            valid_values = section_values.dropna()
            if valid_values.empty:
                continue

            sorted_unique = sorted(valid_values.unique())
            first_min = float(sorted_unique[0])
            second_min = float(sorted_unique[1]) if len(sorted_unique) > 1 else None
            col_idx = df.columns.get_loc(col_name) + 1

            for row_idx, value in enumerate(section_values.tolist(), start=2):
                if pd.notna(value):
                    if abs(float(value) - first_min) < 1e-9:
                        ws.cell(row=row_idx, column=col_idx).fill = orange_fill
                    elif second_min is not None and abs(float(value) - second_min) < 1e-9:
                        ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
    
    def _create_combined_files(self, date_str: str) -> None:
        """Create combined Excel files for different formats."""
        def sort_by_predicted_total(df: pd.DataFrame) -> pd.DataFrame:
            """Sort by 預估總段速 ascending (faster first), keeping missing values at the end."""
            if df.empty or '預估總段速' not in df.columns:
                return df

            sorted_df = df.copy()
            sorted_df['預估總段速數值'] = pd.to_numeric(sorted_df['預估總段速'], errors='coerce')

            sort_cols = ['預估總段速數值']
            ascending = [True]

            # Stable tie-breakers for readability.
            if '檔位' in sorted_df.columns:
                sorted_df['檔位數值'] = pd.to_numeric(sorted_df['檔位'], errors='coerce')
                sort_cols.append('檔位數值')
                ascending.append(True)
            elif '馬匹編號' in sorted_df.columns:
                sorted_df['馬匹編號數值'] = pd.to_numeric(sorted_df['馬匹編號'], errors='coerce')
                sort_cols.append('馬匹編號數值')
                ascending.append(True)

            sorted_df = sorted_df.sort_values(by=sort_cols, ascending=ascending, na_position='last')

            drop_cols = ['預估總段速數值', '檔位數值', '馬匹編號數值']
            existing_drop_cols = [col for col in drop_cols if col in sorted_df.columns]
            return sorted_df.drop(columns=existing_drop_cols)

        # Standard combined file
        standard_sheets_created = 0
        with pd.ExcelWriter(f"Racecard_{date_str}.xlsx", engine='openpyxl') as writer:
            for race_no in range(1, 12):
                try:
                    df = pd.read_excel(f"Racecard_{date_str}_{race_no}.xlsx")
                    if not df.empty:
                        df = sort_by_predicted_total(df)
                        sheet_name = f"Race_{race_no}"
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        self._highlight_fastest_section_cells(writer, sheet_name, df)
                        self._apply_fallback_italic(writer, sheet_name, df)
                        standard_sheets_created += 1
                except FileNotFoundError:
                    logger.warning(f"File Racecard_{date_str}_{race_no}.xlsx not found")
                except Exception as e:
                    logger.error(f"Error adding Race {race_no} to standard file: {e}")
        
        if standard_sheets_created == 0:
            logger.warning(f"No sheets created for standard file Racecard_{date_str}.xlsx")
        
        # Terry format
        terry_sheets_created = 0
        with pd.ExcelWriter(f"Racecard_{date_str}_Terry.xlsx", engine='openpyxl') as writer:
            for race_no in range(1, 12):
                try:
                    df = pd.read_excel(f"Racecard_{date_str}_{race_no}.xlsx")
                    if not df.empty:
                        df = sort_by_predicted_total(df)
                        df_terry = self._create_terry_sheet(df)
                        if not df_terry.empty:
                            sheet_name = f"Race_{race_no}"
                            df_terry.to_excel(writer, sheet_name=sheet_name, index=False)
                            self._highlight_fastest_section_cells(writer, sheet_name, df_terry)
                            self._apply_fallback_italic(writer, sheet_name, df_terry)
                            terry_sheets_created += 1
                except FileNotFoundError:
                    logger.warning(f"File Racecard_{date_str}_{race_no}.xlsx not found for Terry format")
                except Exception as e:
                    logger.error(f"Error creating Terry sheet for Race {race_no}: {e}")
            
            # Create a dummy sheet if no sheets were created
            if terry_sheets_created == 0:
                dummy_df = pd.DataFrame({'Message': ['No race data available']})
                dummy_df.to_excel(writer, sheet_name="No_Data", index=False)
                logger.warning("Created dummy sheet for Terry format due to no valid data")
        
        logger.info(f"Created combined files: Standard({standard_sheets_created} sheets), Terry({terry_sheets_created} sheets)")

    def _create_positioning_file(self, date_str: str) -> None:
        """Create a dedicated Excel file for simulated running positions and generate visualization plots."""
        positioning_columns = [
            '場次', '馬匹編號', '馬名', '評分', '負磅', '檔位', '騎師', '練馬師', '騎練合作上名率',
            '近3仗樣本數', '近況來源',
            '預估今場第 1 段', '預估今場第 2 段', '預估今場第 3 段',
            '預估今場第 4 段', '預估今場第 5 段', '預估今場第 6 段',
            '預估頭段(第1+2段)', '預估末段(最後2段)', '預估總段速', '勝出機率', '公平賠率(扣稅)',
            '預計各分段排名', '模擬走勢', '模擬走位'
        ]

        def build_prediction_dataframe(prediction_map: Dict[int, pd.DataFrame]) -> pd.DataFrame:
            """Build prediction table with one row per race and only top 4 by 預估總段速."""
            if not prediction_map:
                return pd.DataFrame()

            race_nos = sorted(prediction_map.keys())
            rows: List[Dict[str, str]] = []

            for race_no in race_nos:
                race_df = prediction_map[race_no].head(4).copy()
                horse_nos = race_df['馬匹編號'].astype(str).tolist()
                horse_names = race_df['馬名'].fillna('').astype(str).tolist()
                sources = race_df['近況來源'].fillna('').astype(str).tolist() if '近況來源' in race_df.columns else []
                top_win_prob = ''
                top_fair_odds = ''
                if '勝出機率' in race_df.columns and len(race_df) > 0:
                    top_prob_val = pd.to_numeric(race_df.iloc[0]['勝出機率'], errors='coerce')
                    if pd.notna(top_prob_val):
                        top_win_prob = f"{float(top_prob_val):.2f}%"
                if '公平賠率(扣稅)' in race_df.columns and len(race_df) > 0:
                    top_fair_odds_val = pd.to_numeric(race_df.iloc[0]['公平賠率(扣稅)'], errors='coerce')
                    if pd.notna(top_fair_odds_val):
                        top_fair_odds = f"{float(top_fair_odds_val):.2f}"

                while len(horse_nos) < 4:
                    horse_nos.append('')
                    horse_names.append('')
                while len(sources) < 4:
                    sources.append('')

                row: Dict[str, str] = {'場次': f'第{race_no}場'}
                for rank_idx in range(4):
                    row[f'第{rank_idx + 1}名馬號'] = horse_nos[rank_idx]
                    row[f'第{rank_idx + 1}名馬名'] = horse_names[rank_idx]
                row['第1名勝出機率'] = top_win_prob
                row['第1名公平賠率(扣稅)'] = top_fair_odds
                # Source columns go LAST so the visible col % 2 pattern is undisturbed.
                for rank_idx in range(4):
                    row[f'_src_{rank_idx + 1}'] = sources[rank_idx]
                rows.append(row)

            return pd.DataFrame(rows)

        def build_fastest_800_dataframe(fastest_map: Dict[int, pd.DataFrame]) -> pd.DataFrame:
            """Build one-row-per-race summary for fastest head 800m and tail 800m horses."""
            if not fastest_map:
                return pd.DataFrame()

            rows: List[Dict[str, str]] = []
            race_nos = sorted(fastest_map.keys())

            for race_no in race_nos:
                race_df = fastest_map[race_no].copy()
                race_df['馬匹編號數值'] = pd.to_numeric(race_df.get('馬匹編號'), errors='coerce')
                race_df['檔位數值'] = pd.to_numeric(race_df.get('檔位'), errors='coerce')

                row: Dict[str, str] = {
                    '場次': f'第{race_no}場',
                    '頭800最快馬號': '',
                    '頭800最快馬名': '',
                    '頭800時間': '',
                    '尾800最快馬號': '',
                    '尾800最快馬名': '',
                    '尾800時間': ''
                }

                if '預估頭段(第1+2段)' in race_df.columns:
                    head_df = race_df.dropna(subset=['預估頭段(第1+2段)']).sort_values(
                        by=['預估頭段(第1+2段)', '檔位數值', '馬匹編號數值'],
                        ascending=[True, True, True],
                        na_position='last'
                    )
                    if not head_df.empty:
                        head_top = head_df.iloc[0]
                        row['頭800最快馬號'] = str(head_top.get('馬匹編號', ''))
                        row['頭800最快馬名'] = str(head_top.get('馬名', ''))
                        row['頭800時間'] = f"{float(head_top['預估頭段(第1+2段)']):.3f}"

                if '預估末段(最後2段)' in race_df.columns:
                    tail_df = race_df.dropna(subset=['預估末段(最後2段)']).sort_values(
                        by=['預估末段(最後2段)', '檔位數值', '馬匹編號數值'],
                        ascending=[True, True, True],
                        na_position='last'
                    )
                    if not tail_df.empty:
                        tail_top = tail_df.iloc[0]
                        row['尾800最快馬號'] = str(tail_top.get('馬匹編號', ''))
                        row['尾800最快馬名'] = str(tail_top.get('馬名', ''))
                        row['尾800時間'] = f"{float(tail_top['預估末段(最後2段)']):.3f}"

                rows.append(row)

            return pd.DataFrame(rows)

        def style_prediction_sheet(writer: pd.ExcelWriter, max_row: int, max_col: int) -> None:
            """Apply reference-like formatting (yellow/gray blocks) for prediction sheet."""
            ws = writer.book['prediction']

            yellow_fill = PatternFill(fill_type='solid', fgColor='FFC000')
            gray_fill = PatternFill(fill_type='solid', fgColor='E6E6E6')
            header_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
            thin_side = Side(style='thin', color='BFBFBF')
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

            # Header style
            for col in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # Body style: race gray; horse no yellow; horse name gray.
            for row in range(2, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if col == 1:
                        cell.fill = gray_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col == max_col:
                        cell.fill = gray_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col % 2 == 0:
                        cell.fill = yellow_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.fill = gray_fill
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = thin_border

            ws.freeze_panes = 'A2'
            ws.column_dimensions['A'].width = 6
            for col in range(2, max_col + 1):
                col_letter = ws.cell(row=1, column=col).column_letter
                if col % 2 == 0:
                    ws.column_dimensions[col_letter].width = 10
                else:
                    ws.column_dimensions[col_letter].width = 14

        def style_fastest_800_sheet(writer: pd.ExcelWriter, max_row: int, max_col: int) -> None:
            """Apply simple tabular formatting for fastest 800 summary sheet."""
            ws = writer.book['fastest_800']

            header_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
            body_fill = PatternFill(fill_type='solid', fgColor='FFF2CC')
            thin_side = Side(style='thin', color='BFBFBF')
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

            for col in range(1, max_col + 1):
                header_cell = ws.cell(row=1, column=col)
                header_cell.fill = header_fill
                header_cell.font = Font(bold=True)
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
                header_cell.border = thin_border

            for row in range(2, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = body_fill
                    cell.border = thin_border
                    if col in [1, 2, 4, 5, 7]:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

            ws.freeze_panes = 'A2'
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 16
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 16
            ws.column_dimensions['G'].width = 10

        sheets_created = 0
        prediction_data: Dict[int, pd.DataFrame] = {}
        fastest_800_data: Dict[int, pd.DataFrame] = {}
        with pd.ExcelWriter(f"Racecard_{date_str}_走位模擬.xlsx", engine='openpyxl') as writer:
            for race_no in range(1, 12):
                try:
                    df = pd.read_excel(f"Racecard_{date_str}_{race_no}.xlsx")
                    if df.empty:
                        continue

                    existing_columns = [col for col in positioning_columns if col in df.columns]
                    if not existing_columns:
                        continue

                    df_positioning = df[existing_columns].copy()
                    if '預估總段速' in df_positioning.columns:
                        df_positioning['預估總段速數值'] = pd.to_numeric(df_positioning['預估總段速'], errors='coerce')
                        if '檔位' in df_positioning.columns:
                            df_positioning['檔位數值'] = pd.to_numeric(df_positioning['檔位'], errors='coerce')
                            df_positioning = df_positioning.sort_values(
                                by=['預估總段速數值', '檔位數值'],
                                ascending=[True, True],
                                na_position='last'
                            )
                            df_positioning = df_positioning.drop(columns=['檔位數值'])
                        else:
                            df_positioning = df_positioning.sort_values(
                                by=['預估總段速數值'],
                                ascending=[True],
                                na_position='last'
                            )
                        df_positioning = df_positioning.drop(columns=['預估總段速數值'])

                    # Build prediction source sorted by 預估總段速 (smaller = faster).
                    if {'馬匹編號', '馬名', '預估總段速'}.issubset(df.columns):
                        pred_df = df[['馬匹編號', '馬名', '檔位', '預估總段速']].copy()
                        if '近況來源' in df.columns:
                            pred_df['近況來源'] = df['近況來源'].values
                        if '勝出機率' in df.columns:
                            pred_df['勝出機率'] = pd.to_numeric(df['勝出機率'], errors='coerce')
                        if '公平賠率(扣稅)' in df.columns:
                            pred_df['公平賠率(扣稅)'] = pd.to_numeric(df['公平賠率(扣稅)'], errors='coerce')
                        pred_df['檔位數值'] = pd.to_numeric(pred_df.get('檔位'), errors='coerce')
                        pred_df['預估總段速數值'] = pd.to_numeric(pred_df['預估總段速'], errors='coerce')
                        pred_df['缺值'] = pred_df['預估總段速數值'].isna()
                        pred_df = pred_df.sort_values(
                            by=['缺值', '預估總段速數值', '檔位數值'],
                            ascending=[True, True, True],
                            na_position='last'
                        )
                        keep_cols = ['馬匹編號', '馬名', '預估總段速數值']
                        if '勝出機率' in pred_df.columns:
                            keep_cols.append('勝出機率')
                        if '公平賠率(扣稅)' in pred_df.columns:
                            keep_cols.append('公平賠率(扣稅)')
                        if '近況來源' in pred_df.columns:
                            keep_cols.append('近況來源')
                        prediction_data[race_no] = pred_df[keep_cols].reset_index(drop=True)

                    if {'馬匹編號', '馬名'}.issubset(df.columns):
                        fastest_cols = ['馬匹編號', '馬名']
                        if '檔位' in df.columns:
                            fastest_cols.append('檔位')
                        if '預估頭段(第1+2段)' in df.columns:
                            fastest_cols.append('預估頭段(第1+2段)')
                        if '預估末段(最後2段)' in df.columns:
                            fastest_cols.append('預估末段(最後2段)')

                        if ('預估頭段(第1+2段)' in fastest_cols) or ('預估末段(最後2段)' in fastest_cols):
                            fastest_df = df[fastest_cols].copy()
                            if '預估頭段(第1+2段)' in fastest_df.columns:
                                fastest_df['預估頭段(第1+2段)'] = pd.to_numeric(
                                    fastest_df['預估頭段(第1+2段)'], errors='coerce'
                                )
                            if '預估末段(最後2段)' in fastest_df.columns:
                                fastest_df['預估末段(最後2段)'] = pd.to_numeric(
                                    fastest_df['預估末段(最後2段)'], errors='coerce'
                                )
                            fastest_800_data[race_no] = fastest_df

                    sheet_name = f"Race_{race_no}"
                    df_positioning.to_excel(writer, sheet_name=sheet_name, index=False)
                    self._highlight_fastest_section_cells(writer, sheet_name, df_positioning)
                    self._apply_fallback_italic(writer, sheet_name, df_positioning)
                    sheets_created += 1
                except FileNotFoundError:
                    logger.warning(f"File Racecard_{date_str}_{race_no}.xlsx not found for positioning format")
                except Exception as e:
                    logger.error(f"Error creating positioning sheet for Race {race_no}: {e}")

            if sheets_created == 0:
                dummy_df = pd.DataFrame({'Message': ['No race positioning data available']})
                dummy_df.to_excel(writer, sheet_name="No_Data", index=False)
                logger.warning("Created dummy sheet for positioning format due to no valid data")
            elif prediction_data:
                prediction_sheet_df = build_prediction_dataframe(prediction_data)
                prediction_sheet_df.to_excel(writer, sheet_name='prediction', index=False)
                # Visible column count excludes hidden source-tracking columns (_src_N).
                visible_col_count = sum(1 for c in prediction_sheet_df.columns if not c.startswith('_src_'))
                style_prediction_sheet(
                    writer,
                    max_row=prediction_sheet_df.shape[0] + 1,
                    max_col=visible_col_count,
                )
                # Italic horse names for cross-distance / cross-venue fallback horses.
                ws_pred = writer.book['prediction']
                for rank_idx in range(1, 5):
                    src_col = f'_src_{rank_idx}'
                    name_col = f'第{rank_idx}名馬名'
                    if src_col not in prediction_sheet_df.columns or name_col not in prediction_sheet_df.columns:
                        continue
                    name_col_idx = prediction_sheet_df.columns.get_loc(name_col) + 1
                    for row_data_idx, src_val in enumerate(prediction_sheet_df[src_col].tolist(), start=2):
                        if str(src_val) not in ('同場同程', '', 'nan'):
                            cell = ws_pred.cell(row=row_data_idx, column=name_col_idx)
                            ef = cell.font
                            cell.font = Font(name=ef.name, size=ef.size, bold=ef.bold, italic=True, color=ef.color)
                # Hide source-tracking columns from the sheet.
                for rank_idx in range(1, 5):
                    src_col = f'_src_{rank_idx}'
                    if src_col in prediction_sheet_df.columns:
                        src_col_idx = prediction_sheet_df.columns.get_loc(src_col) + 1
                        col_letter = ws_pred.cell(row=1, column=src_col_idx).column_letter
                        ws_pred.column_dimensions[col_letter].hidden = True

                if fastest_800_data:
                    fastest_800_sheet_df = build_fastest_800_dataframe(fastest_800_data)
                    if not fastest_800_sheet_df.empty:
                        fastest_800_sheet_df.to_excel(writer, sheet_name='fastest_800', index=False)
                        style_fastest_800_sheet(
                            writer,
                            max_row=fastest_800_sheet_df.shape[0] + 1,
                            max_col=fastest_800_sheet_df.shape[1]
                        )

        logger.info(f"Created positioning file: Racecard_{date_str}_走位模擬.xlsx ({sheets_created} sheets)")
    
    def download_all_racecards(self) -> None:
        """Download all racecards for the configured race date and course."""
        date_str = self.race_date.strftime('%Y%m%d')
        logger.info(f"Starting racecard download for {self.race_date} at {self.race_course}")
        
        successful_downloads = 0
        individual_files = []  # Track individual files created
        
        # Download individual race cards (typically 1-11 races)
        for race_no in range(1, 12):
            logger.info(f"Processing {self.race_date} {self.race_course} Race {race_no}")
            
            df = self._scrape_race_data(race_no)
            
            if df is not None and not df.empty:
                # Save individual race file
                filename = f"Racecard_{date_str}_{race_no}.xlsx"
                try:
                    df.to_excel(filename, index=False)
                    logger.info(f"Saved {filename}")
                    successful_downloads += 1
                    individual_files.append(filename)  # Track the file
                except Exception as e:
                    logger.error(f"Error saving {filename}: {e}")
            else:
                logger.warning(f"No data found for Race {race_no}")
        
        # Create combined files in different formats
        if successful_downloads > 0:
            try:
                self._create_combined_files(date_str)
                self._create_positioning_file(date_str)
                
                # Remove individual race files after successful combined file creation
                self._cleanup_individual_files(individual_files)
                
            except Exception as e:
                logger.error(f"Error creating combined files: {e}")
        else:
            logger.warning("No successful downloads, skipping combined file creation")
        
        logger.info(f"Racecard download completed. Successfully processed {successful_downloads} races.")
    
    def _cleanup_individual_files(self, individual_files: List[str]) -> None:
        """Remove individual race files after combined files are created."""
        removed_count = 0
        
        for filename in individual_files:
            try:
                file_path = Path(filename)
                if file_path.exists():
                    file_path.unlink()  # Delete the file
                    logger.info(f"Removed individual file: {filename}")
                    removed_count += 1
                else:
                    logger.warning(f"File not found for cleanup: {filename}")
            except Exception as e:
                logger.error(f"Error removing file {filename}: {e}")
        
        logger.info(f"Cleanup completed. Removed {removed_count} individual race files.")
    
def main():
    # Configuration
    race_date = date(2026, 5, 27)
    race_course = "HV"  # ST / HV
    
    # Create downloader and process
    downloader = HKJCRacecardDownloader(race_date, race_course)
    downloader.download_all_racecards()
    
    logger.info("Racecard download completed!")

if __name__ == "__main__":
    main()